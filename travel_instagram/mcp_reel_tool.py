from __future__ import annotations

import json
import logging
import csv
import os
import random
import re
import string
import subprocess
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Literal

import httpx
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont

from travel_instagram import config
from travel_instagram import media_processor

logger = logging.getLogger(__name__)


REEL_TYPE = Literal["image", "video"]


_DESTINATION_KEYWORDS = (
    "destination",
    "to",
    "fly to",
    "fly",
    "going to",
    "visit",
    "travel to",
)


def _slugify(s: str, max_len: int = 40) -> str:
    s = s.strip().lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    s = s.strip("-")
    return (s[:max_len] or "reel") + "-" + datetime.now(timezone.utc).strftime("%H%M%S")


def _title_case(s: str) -> str:
    # Preserve things like "New York" nicely; avoids lowercasing acronyms.
    s = re.sub(r"\s+", " ", s.strip())
    return s[:1].upper() + s[1:]


def _segment_fly_label(row: PriceRow, fallback_destination: str) -> str:
    """
    Prefer explicit city + country from price rows, e.g. "Fly Berlin, Germany".
    Falls back to parsed destination when row fields are missing.
    """
    city = str(getattr(row, "destination", "") or "").strip()
    country = str(getattr(row, "country", "") or "").strip()
    c_abbr = _country_abbr(country)
    if city and country:
        lc = city.lower()
        if country.lower() in lc:
            return f"Fly to {city}"
        return f"Fly to {city}, {c_abbr or country}"
    if city:
        return f"Fly to {city}"
    return f"Fly to {fallback_destination.strip()}"


def _country_abbr(country: str) -> str | None:
    c = (country or "").strip()
    if not c:
        return None
    c_low = c.lower()
    known = {
        "united kingdom": "UK",
        "united states": "US",
        "united states of america": "US",
        "india": "IN",
        "germany": "DE",
        "france": "FR",
        "spain": "ES",
        "italy": "IT",
        "portugal": "PT",
        "ireland": "IE",
        "netherlands": "NL",
        "belgium": "BE",
        "poland": "PL",
        "denmark": "DK",
        "austria": "AT",
        "switzerland": "CH",
        "turkey": "TR",
        "greece": "GR",
        "cyprus": "CY",
        "czech republic": "CZ",
        "czechia": "CZ",
    }
    if c_low in known:
        return known[c_low]
    # Fallback: initials for multi-word names.
    parts = [p for p in re.split(r"[\s\-]+", c) if p]
    if len(parts) >= 2:
        ini = "".join(p[0] for p in parts[:2]).upper()
        if len(ini) == 2 and ini.isalpha():
            return ini
    return None


def _normalize_price_text(price_text: str | None) -> str | None:
    if price_text is None:
        return None
    s = str(price_text).strip()
    if not s:
        return None
    # Common mojibake in CSV exports.
    s = s.replace("�", "€").replace("?", "€")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _parse_intent_from_prompt(prompt: str) -> dict[str, Any]:
    p = prompt.lower()
    intent = "travel"
    if any(w in p for w in ("price", "cheapest", "cheap", "from €", "eur", "euro")):
        intent = "pricing"

    mode: str | None = None
    if "airport" in p:
        mode = "airport"
    elif "walk" in p or "walking" in p:
        mode = "walking"
    elif "cloud" in p or "flying" in p or "fly" in p or "airplane" in p:
        mode = "flight"
    elif "train" in p:
        mode = "train"
    elif "car" in p or "road trip" in p:
        mode = "road"

    return {"intent": intent, "mode": mode}


def _extract_destination(prompt: str) -> str | None:
    """
    Heuristic parsing:
    - "destination Paris" -> Paris
    - "fly to Paris" -> Paris
    - "Beautiful destination Paris, now fly at" -> Paris
    - Otherwise: first capitalized phrase of 1–3 words.
    """
    p = prompt.strip()
    m = re.search(r"destination\s+([A-Za-z][A-Za-z\s\-]{1,40})", p, flags=re.IGNORECASE)
    if m:
        return _title_case(m.group(1).split(",")[0])

    m = re.search(r"fly\s+to\s+([A-Za-z][A-Za-z\s\-]{1,40})", p, flags=re.IGNORECASE)
    if m:
        return _title_case(m.group(1).split(",")[0])

    m = re.search(r"\bto\s+([A-Za-z][A-Za-z\s\-]{1,40})\b", p, flags=re.IGNORECASE)
    if m:
        # Avoid "to" in "... ready to ..." by filtering common non-destination words.
        candidate = m.group(1).strip()
        if candidate.lower() in {"you", "them", "the", "reach", "airport"}:
            return None
        return _title_case(candidate.split(",")[0])

    # Fallback: capitalized 1–3 words.
    cap = re.findall(r"\b([A-Z][a-z]+(?:\s+[A-Z][a-z]+){0,2})\b", p)
    if cap:
        return cap[0].strip()
    return None


def _extract_origin(prompt: str) -> str | None:
    p = prompt.strip()
    m = re.search(r"\bfrom\s+([A-Za-z][A-Za-z\s\-]{1,40})", p, flags=re.IGNORECASE)
    if not m:
        m = re.search(r"\bat\s+([A-Za-z][A-Za-z\s\-]{1,40})", p, flags=re.IGNORECASE)
    if not m:
        return None
    candidate = m.group(1).split(",")[0].strip()
    if candidate.lower() in {"europe", "world"}:
        return None
    return _title_case(candidate)


def parse_prompt(prompt: str) -> dict[str, Any]:
    if not prompt or not str(prompt).strip():
        raise ValueError("Prompt is required.")
    dest = _extract_destination(prompt)
    if not dest:
        raise ValueError("Could not extract destination from prompt. Example: 'destination Paris'.")
    origin = _extract_origin(prompt)
    intent = _parse_intent_from_prompt(prompt)
    return {"destination": dest, "origin": origin, **intent}


def _norm_header(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def _parse_float_from_any(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    # Extract first number like "55" or "€55"
    m = re.search(r"(\d+(?:[.,]\d+)?)", s)
    if not m:
        return None
    return float(m.group(1).replace(",", "."))


@dataclass(frozen=True)
class PriceRow:
    country: str
    origin: str
    destination: str
    price: float | None
    price_range: str | None
    price_text: str | None


def load_price_rows(excel_path: Path, *, sheet_name: str | None = None) -> list[PriceRow]:
    """
    Load destination price rows from either:
    - Excel: .xlsx/.xls (openpyxl)
    - CSV: .csv (csv.DictReader)
    """
    if not excel_path.is_file():
        raise RuntimeError(f"Prices file not found: {excel_path}")

    if excel_path.suffix.lower() == ".csv":
        return _load_price_rows_csv(excel_path)

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            header_row = row
            break
    if not header_row:
        raise RuntimeError("Excel has no header row.")

    headers = [_norm_header(str(c)) for c in header_row]

    def col(*names: str) -> int | None:
        want = {_norm_header(n) for n in names}
        for idx, h in enumerate(headers):
            if h in want:
                return idx
        return None

    c_country = col("country")
    c_origin = col("origin")
    c_dest = col("destination")
    c_price = col("price", "price_eur", "price (eur)")
    c_price_range = col("price_range", "price range")
    c_price_text = col("price_text", "price text", "price")

    if c_country is None or c_origin is None or c_dest is None:
        raise RuntimeError(
            "Excel must contain columns: country, origin, destination. "
            f"Found headers: {headers}"
        )

    out: list[PriceRow] = []
    for ridx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if ridx == 1:
            continue
        if not row:
            continue
        destination = str(row[c_dest]).strip() if c_dest is not None else ""
        if not destination:
            continue

        country = str(row[c_country]).strip() if c_country is not None else ""
        origin = str(row[c_origin]).strip() if c_origin is not None else ""
        price = _parse_float_from_any(row[c_price]) if c_price is not None else None
        price_range = str(row[c_price_range]).strip() if c_price_range is not None and row[c_price_range] else None
        price_text = str(row[c_price_text]).strip() if c_price_text is not None and row[c_price_text] else None

        out.append(
            PriceRow(
                country=country,
                origin=origin,
                destination=destination,
                price=price,
                price_range=price_range,
                price_text=price_text,
            )
        )
    return out


def _load_price_rows_csv(csv_path: Path) -> list[PriceRow]:
    import io

    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        sample = f.read(8192)
        f.seek(0)

        # Try to sniff delimiters; fall back to comma.
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=[",", ";", "\t", "|"])
        except csv.Error:
            dialect = csv.excel

        reader = csv.DictReader(f, dialect=dialect)
        if not reader.fieldnames:
            raise RuntimeError("CSV has no header row.")

        raw_headers = [str(h or "").strip() for h in reader.fieldnames]
        headers_norm = [_norm_header(h) for h in raw_headers]
        header_norm_to_raw: dict[str, str] = {}
        for raw, nn in zip(raw_headers, headers_norm):
            if nn and nn not in header_norm_to_raw:
                header_norm_to_raw[nn] = raw

        def pick(*names: str) -> str | None:
            want = {_norm_header(n) for n in names}
            for nn, raw in header_norm_to_raw.items():
                if nn in want:
                    return raw
            return None

        c_country = pick("country")
        c_origin = pick("origin")
        c_dest = pick("destination")
        c_price = pick("price", "price_eur", "price (eur)")
        c_price_range = pick("price_range", "price range")
        c_price_text = pick("price_text", "price text")

        if c_country is None or c_origin is None or c_dest is None:
            raise RuntimeError(
                "CSV must contain columns: country, origin, destination. "
                f"Found headers: {raw_headers}"
            )

        out: list[PriceRow] = []
        for row in reader:
            if not row:
                continue
            destination = str(row.get(c_dest) or "").strip()
            if not destination:
                continue

            country = str(row.get(c_country) or "").strip()
            origin = str(row.get(c_origin) or "").strip()

            price = _parse_float_from_any(row.get(c_price)) if c_price else None
            price_range = str(row.get(c_price_range) or "").strip() if c_price_range else None
            if price_range == "":
                price_range = None
            price_text = str(row.get(c_price_text) or "").strip() if c_price_text else None
            if price_text == "":
                price_text = None

            out.append(
                PriceRow(
                    country=country,
                    origin=origin,
                    destination=destination,
                    price=price,
                    price_range=price_range,
                    price_text=price_text,
                )
            )
        return out


def select_price_rows_for_destination(
    rows: list[PriceRow],
    destination_query: str,
    *,
    count: tuple[int, int] = (4, 5),
) -> list[PriceRow]:
    # Clean up punctuation so "Ireland," still matches.
    want = destination_query.strip().lower()
    want = re.sub(r"[^a-z0-9\s\-]", " ", want)
    want = re.sub(r"\s+", " ", want).strip()
    matches: list[PriceRow] = []
    for r in rows:
        d = (r.destination or "").strip().lower()
        if not d:
            continue
        d_clean = re.sub(r"[^a-z0-9\s\-]", " ", d)
        d_clean = re.sub(r"\s+", " ", d_clean).strip()
        if d_clean == want or want in d_clean or d_clean in want:
            matches.append(r)

    # If the query looks like a country (e.g. "Ireland") but `destination` column
    # contains cities, fall back to matching `country`.
    if not matches:
        country_matches: list[PriceRow] = []
        for r in rows:
            c = (r.country or "").strip().lower()
            if not c:
                continue
            c_clean = re.sub(r"[^a-z0-9\s\-]", " ", c)
            c_clean = re.sub(r"\s+", " ", c_clean).strip()
            if c_clean == want or want in c_clean or c_clean in want:
                country_matches.append(r)
        if country_matches:
            matches = country_matches
        else:
            raise RuntimeError(
                f"No price rows match destination={destination_query!r} (checked destination and country columns)"
            )

    # Choose the cheapest first; then random pick within top.
    matches.sort(key=lambda r: (r.price if r.price is not None else 1e18))
    top_n = min(len(matches), 10)
    top = matches[:top_n]

    k_target = random.randint(count[0], count[1])
    if len(top) >= k_target:
        chosen = random.sample(top, k_target)
    else:
        # If the dataset has only 1–2 matching destinations (common with small CSV extracts),
        # pad by sampling with replacement so the reel still has 4–5 segments.
        chosen = [random.choice(top) for _ in range(k_target)]

    # Keep overall order roughly cheapest->expensive for nicer progression.
    chosen.sort(key=lambda r: (r.price if r.price is not None else 1e18))
    return chosen


def _pexels_headers() -> dict[str, str]:
    if not config.PEXELS_API_KEY:
        raise RuntimeError("PEXELS_API_KEY is not set.")
    return {"Authorization": config.PEXELS_API_KEY}


def _best_photo_src_url(photo: dict[str, Any]) -> str | None:
    src = photo.get("src") or {}
    for key in ("original", "large2x", "portrait", "large", "medium"):
        u = src.get(key)
        if u:
            return str(u)
    return None


_TRANSPORT_RE = re.compile(
    r"\b(airplane|aeroplane|plane|flight|flying|airport|runway|boarding|takeoff|take-off|landing|cloud|clouds|sky|pilot|cockpit|luggage)\b",
    flags=re.IGNORECASE,
)


def _looks_transport_asset(blob: str | None) -> bool:
    if not blob:
        return False
    return _TRANSPORT_RE.search(blob) is not None


def _score_video_file(file_obj: dict[str, Any], *, vid_obj: dict[str, Any]) -> float:
    w = int(file_obj.get("width") or 0)
    h = int(file_obj.get("height") or 0)
    if w < 1 or h < 1:
        return -1e9
    if h < w * 0.9:
        return -1e9
    q = str(file_obj.get("quality") or "").lower()
    q_bonus = 2.0 if q == "hd" else 1.0 if q == "sd" else 0.5
    duration = float(vid_obj.get("duration") or 0)
    return h * w * q_bonus + (h / max(1, w)) * 1000 + duration


def _best_portrait_video_link(videos: list[dict[str, Any]], *, exclude_urls: set[str] | None = None) -> dict[str, Any] | None:
    """
    Choose the first portrait-ish clip by Pexels relevance ordering.

    We intentionally avoid resolution scoring here because it can drift away from
    the search intent and select less-relevant but larger clips.
    """
    exclude_urls = exclude_urls or set()
    for vid in videos:
        meta_blob = " ".join(
            [
                str(vid.get("url") or ""),
                str(vid.get("image") or ""),
                str((vid.get("user") or {}).get("name") or ""),
            ]
        )
        for f in vid.get("video_files") or []:
            link = f.get("link")
            if not link or str(link) in exclude_urls:
                continue
            if _looks_transport_asset(f"{link} {meta_blob}"):
                continue
            w = int(f.get("width") or 0)
            h = int(f.get("height") or 0)
            if w < 1 or h < 1:
                continue
            if h < w * 0.9:
                continue
            return {
                "url": str(link),
                "width": w,
                "height": h,
                "duration": float(vid.get("duration") or 0),
            }
    return None


def build_pexels_query(
    destination: str,
    origin: str | None,
    *,
    mode: str | None,
    extra_terms: str | None = None,
) -> str:
    d = destination.strip()
    o = (origin or "").strip()

    # Always include destination-specific visual keywords; otherwise "flight"
    # / "airport" terms can dominate and return generic footage.
    base_terms = "landmarks attractions city skyline scenic view"

    # Destination-first visual query. Do not include transport / flight hints here.
    mode_terms = "city landmarks attractions travel"

    extra = (extra_terms or "").strip()
    if extra:
        base_terms = extra + " " + base_terms

    if o:
        return f"{d} {base_terms} {o} {mode_terms} portrait"
    return f"{d} {base_terms} {mode_terms} portrait"


def _download_pexels_image(url: str, dest: Path, *, client: httpx.Client) -> None:
    media_processor.download_binary(url, dest, client=client)


def _download_pexels_video(url: str, dest: Path, *, client: httpx.Client) -> None:
    media_processor.download_binary(url, dest, client=client)


def fetch_media_for_destination_entry(
    *,
    destination: str,
    origin: str | None,
    price_text: str | None,
    mode: str | None,
    image_keywords: str | None,
    video_keywords: str | None,
    work_dir: Path,
    used_image_urls: set[str],
    used_video_urls: set[str],
    client: httpx.Client,
) -> dict[str, Any]:
    """
    Fetch 1 image (always) and optionally 1 video. Returns local file paths.
    """
    wdir = work_dir
    wdir.mkdir(parents=True, exist_ok=True)
    headers = _pexels_headers()
    image_query = build_pexels_query(
        destination,
        origin,
        mode=mode,
        extra_terms=image_keywords,
    )
    video_query = build_pexels_query(
        destination,
        origin,
        mode=mode,
        extra_terms=video_keywords,
    )
    # Strict destination-only query ladder (no flight/sky transport fallback).
    fallback_img_queries = [
        image_query,
        f"{destination} city landmarks attractions portrait",
        f"{destination} things to do popular spots portrait",
        f"{destination} tourism places to visit portrait",
        f"{destination} scenic view nature city portrait",
        f"{destination} architecture old town street portrait",
    ]

    image_path = wdir / f"img_{slugify_for_fs(destination)}.jpg"
    video_path = wdir / f"vid_{slugify_for_fs(destination)}.mp4"

    # Images (try multiple queries so we don't fail when the first search is thin).
    img_url: str | None = None
    last_err: Exception | None = None
    for q in fallback_img_queries:
        try:
            ir = client.get(
                "https://api.pexels.com/v1/search",
                headers=headers,
                params={
                    "query": q,
                    "per_page": 15,
                    "orientation": "portrait",
                },
            )
            ir.raise_for_status()
            idata = ir.json()
            photos = idata.get("photos") or []

            for p in photos:
                u = _best_photo_src_url(p)
                blob = " ".join(
                    [
                        str(u or ""),
                        str(p.get("url") or ""),
                        str(p.get("alt") or ""),
                        str((p.get("src") or {}).get("original") or ""),
                        str((p.get("src") or {}).get("large2x") or ""),
                    ]
                )
                if u and u not in used_image_urls and not _looks_transport_asset(blob):
                    img_url = u
                    break
            if img_url:
                break
            # Fallback to first if none were unique.
            if photos:
                u = _best_photo_src_url(photos[0])
                p0 = photos[0]
                blob0 = " ".join(
                    [
                        str(u or ""),
                        str(p0.get("url") or ""),
                        str(p0.get("alt") or ""),
                        str((p0.get("src") or {}).get("original") or ""),
                        str((p0.get("src") or {}).get("large2x") or ""),
                    ]
                )
                if u and u not in used_image_urls and not _looks_transport_asset(blob0):
                    img_url = u
                    break
        except Exception as e:
            last_err = e

    if not img_url:
        msg = f"No Pexels image found for {destination!r}"
        if last_err:
            msg += f" (last error: {last_err})"
        raise RuntimeError(msg)

    used_image_urls.add(img_url)
    _download_pexels_image(img_url, image_path, client=client)

    # Videos (optional) - also use fallbacks.
    video = None
    fallback_video_queries = [
        video_query,
        f"{destination} cinematic city landmarks vertical",
        f"{destination} things to do attractions walking vertical",
        f"{destination} tourism places to visit vertical",
        f"{destination} scenic city nature drone vertical",
        f"{destination} architecture old town streets vertical",
    ]
    for q in fallback_video_queries:
        try:
            vr = client.get(
                "https://api.pexels.com/videos/search",
                headers=headers,
                params={
                    "query": q,
                    "per_page": 10,
                    "orientation": "portrait",
                },
            )
            vr.raise_for_status()
            vdata = vr.json()
            vids = vdata.get("videos") or []
            video = _best_portrait_video_link(vids, exclude_urls=used_video_urls)
            if video:
                break
        except httpx.HTTPError:
            continue

    if video and video.get("url"):
        used_video_urls.add(video["url"])
        _download_pexels_video(video["url"], video_path, client=client)
        return {
            "image_path": image_path,
            "video_path": video_path,
            "video_meta": video,
            "pexels_query_image": image_query,
            "pexels_query_video": video_query,
        }

    return {
        "image_path": image_path,
        "video_path": None,
        "video_meta": None,
        "pexels_query_image": image_query,
        "pexels_query_video": video_query,
    }


def slugify_for_fs(s: str) -> str:
    s = s.strip().lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    s = s.strip("-")
    return s[:48] or "x"


def _render_segment_overlay_png(
    out_png: Path,
    *,
    destination: str,
    price_text: str | None,
    origin: str | None,
) -> Path:
    """
    Full-frame transparent PNG with a rounded card + big text.
    """
    w, h = config.REEL_SIZE
    out_png.parent.mkdir(parents=True, exist_ok=True)

    card_alpha = 208
    img = Image.new("RGBA", (w, h), (0, 0, 0, 0))
    draw = ImageDraw.Draw(img)

    # Fonts: fall back to default if none found.
    def _try_font(candidates: list[str], size: int) -> ImageFont.ImageFont:
        for p in candidates:
            try:
                return ImageFont.truetype(p, size=size)
            except OSError:
                continue
        return ImageFont.load_default()

    # Revert to stronger visual sizes (user preference), while keeping improved spacing.
    title_size = int(h * 0.058)
    price_size = int(h * 0.082)
    small_size = int(h * 0.034)
    title_font = _try_font(
        [r"C:\Windows\Fonts\segoeuib.ttf", r"C:\Windows\Fonts\arialbd.ttf"],
        title_size,
    )
    price_font = _try_font(
        [r"C:\Windows\Fonts\segoeuib.ttf", r"C:\Windows\Fonts\arialbd.ttf"],
        price_size,
    )
    small_font = _try_font(
        [r"C:\Windows\Fonts\seguisb.ttf", r"C:\Windows\Fonts\arial.ttf"],
        small_size,
    )

    destination = (destination or "").strip()
    price_line = _normalize_price_text(price_text) or ""
    origin_line = (f"From {origin}".strip() if origin else "").strip()

    if not destination:
        destination = "Destination"

    # Compute line breaks.
    def wrap(draw_: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont, max_w: int) -> list[str]:
        if not text:
            return []
        words = text.split()
        lines: list[str] = []
        cur: list[str] = []
        for word in words:
            trial = " ".join(cur + [word]).strip()
            bbox = draw_.textbbox((0, 0), trial, font=font)
            if bbox[2] - bbox[0] <= max_w:
                cur.append(word)
            else:
                if cur:
                    lines.append(" ".join(cur))
                cur = [word]
        if cur:
            lines.append(" ".join(cur))
        return lines[:4]

    max_w = int(w * 0.86)
    title_lines = wrap(draw, destination, title_font, max_w)
    price_lines = wrap(draw, price_line, price_font, max_w) if price_line else []
    small_lines: list[str] = []
    if origin_line:
        small_lines.extend(wrap(draw, origin_line, small_font, max_w))

    # Block size.
    line_gap = int(max(8, h * 0.010))
    group_gap = int(max(12, h * 0.014))
    pad_x = int(max(36, w * 0.068))
    pad_y = int(max(24, h * 0.026))

    def line_h(font: ImageFont.ImageFont, txt: str) -> int:
        if not txt:
            return 0
        bb = draw.textbbox((0, 0), txt, font=font)
        return bb[3] - bb[1]

    block_h = (
        sum(line_h(title_font, ln) for ln in title_lines)
        + (max(0, len(title_lines) - 1) * line_gap)
    )
    if price_lines:
        block_h += group_gap + sum(line_h(price_font, ln) for ln in price_lines) + max(0, len(price_lines) - 1) * line_gap
    if small_lines:
        block_h += group_gap + sum(line_h(small_font, ln) for ln in small_lines) + max(0, len(small_lines) - 1) * line_gap

    rect_w = min(w - 40, max_w + pad_x * 2)
    rect_h = block_h + pad_y * 2
    x0 = (w - rect_w) // 2
    y0 = int(h * 0.08)
    y0 = max(8, min(y0, h - rect_h - 8))

    draw.rounded_rectangle(
        (x0, y0, x0 + rect_w, y0 + rect_h),
        radius=int(min(36, h * 0.03)),
        fill=(8, 12, 22, card_alpha),
    )

    # Text draw.
    cy = y0 + pad_y
    center_x = w // 2

    for ln in title_lines:
        if not ln.strip():
            continue
        bb = draw.textbbox((0, 0), ln, font=title_font)
        tw = bb[2] - bb[0]
        tx = center_x - tw // 2
        # Shadow
        draw.text((tx + 3, cy + 3), ln, font=title_font, fill=(0, 0, 0, 175))
        draw.text((tx, cy), ln, font=title_font, fill=(248, 250, 252, 252))
        cy += line_h(title_font, ln) + line_gap

    if price_lines:
        cy += group_gap
        for ln in price_lines:
            if not ln.strip():
                continue
            bb = draw.textbbox((0, 0), ln, font=price_font)
            tw = bb[2] - bb[0]
            tx = center_x - tw // 2
            # Accent + subtle glow / pulse-like visual emphasis.
            draw.text((tx + 5, cy + 5), ln, font=price_font, fill=(20, 90, 145, 160))
            draw.text((tx + 2, cy + 2), ln, font=price_font, fill=(0, 0, 0, 165))
            draw.text((tx, cy), ln, font=price_font, fill=(92, 196, 255, 255))
            cy += line_h(price_font, ln) + line_gap

    if small_lines:
        cy += group_gap
        for ln in small_lines:
            if not ln.strip():
                continue
            bb = draw.textbbox((0, 0), ln, font=small_font)
            tw = bb[2] - bb[0]
            tx = center_x - tw // 2
            draw.text((tx + 2, cy + 2), ln, font=small_font, fill=(0, 0, 0, 165))
            draw.text((tx, cy), ln, font=small_font, fill=(194, 204, 216, 248))
            cy += line_h(small_font, ln) + line_gap

    img.save(out_png)
    return out_png


def _ffmpeg_exe() -> str:
    return media_processor._ensure_ffmpeg()  # type: ignore[attr-defined]


def _run_ffmpeg(cmd: list[str], context: str) -> None:
    proc = subprocess.run(
        cmd,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        check=False,
    )
    if proc.returncode != 0:
        tail = (proc.stderr or "")[-4000:]
        raise RuntimeError(f"ffmpeg failed ({context}) exit={proc.returncode}. stderr_tail:\n{tail}")


def _normalize_and_encode_video_segment(
    video_src: Path,
    overlay_png: Path | None,
    out_mp4: Path,
    *,
    seconds_each: float,
    context: str,
) -> None:
    """
    Normalize video to reel format and optionally overlay a PNG card.
    Audio is dropped.
    """
    w, h = config.REEL_SIZE
    fps = 30
    exe = _ffmpeg_exe()

    vf = (
        f"scale={w}:{h}:force_original_aspect_ratio=increase,"
        f"crop={w}:{h},format=yuv420p,fps={fps}"
    )

    cmd: list[str] = [exe, "-y", "-i", str(video_src)]

    if overlay_png is not None and overlay_png.is_file():
        cmd.extend(["-i", str(overlay_png)])
        cmd.extend(
            [
                "-filter_complex",
                f"[0:v]{vf}[v];[1:v]format=rgba[ov];[v][ov]overlay=0:0:format=auto",
            ]
        )
    else:
        cmd.extend(["-vf", vf])

    cmd.extend(
        [
            "-an",
            "-c:v",
            "libx264",
            "-pix_fmt",
            "yuv420p",
            "-preset",
            "veryfast",
            "-crf",
            "20",
            "-movflags",
            "+faststart",
            "-t",
            f"{seconds_each:.3f}",
            str(out_mp4),
        ]
    )
    _run_ffmpeg(cmd, context)


def _encode_image_segment(
    image_src: Path,
    overlay_png: Path | None,
    out_mp4: Path,
    *,
    seconds_each: float,
    context: str,
) -> None:
    """
    Encode a single cover-cropped frame (RGB rawvideo) into a short segment.
    If overlay_png is provided, we composite it on the frame using Pillow.
    """
    w, h = config.REEL_SIZE
    fps = 30

    with Image.open(image_src) as im:
        rgb = im.convert("RGB")
        frame = media_processor._cover_crop(rgb, (w, h))  # type: ignore[attr-defined]
        if overlay_png is not None and overlay_png.is_file():
            # Composite overlay card onto frame.
            with Image.open(overlay_png) as ov:
                ov_rgba = ov.convert("RGBA")
            base_rgba = frame.convert("RGBA")
            base_rgba.alpha_composite(ov_rgba, (0, 0))
            frame = base_rgba.convert("RGB")

        still_bytes = frame.tobytes()

    media_processor._encode_reel_rawvideo_to_mp4(  # type: ignore[attr-defined]
        [still_bytes],
        seconds_each,
        out_mp4,
        context,
    )


def generate_travel_reel_from_prompt(
    prompt: str,
    *,
    excel_path: str | None = None,
    output_dir: str | None = None,
    music_path: str | None = None,
    search_destination: str | None = None,
    search_origin: str | None = None,
    search_mode: str | None = None,
    image_keywords: str | None = None,
    video_keywords: str | None = None,
    include_music: bool | None = None,
    target_total_segments: int | None = None,
) -> dict[str, Any]:
    """
    Main entry point for the MCP tool.

    Returns JSON-serializable dict with output path and metadata.
    """
    parsed = parse_prompt(prompt)
    destination = search_destination or parsed["destination"]
    origin = search_origin if search_origin is not None else parsed.get("origin")
    mode = search_mode or parsed.get("mode")

    if image_keywords is None:
        image_keywords = "landmarks attractions city skyline scenic view"
    if video_keywords is None:
        video_keywords = "cinematic travel motion vertical"

    def _strip_proper_nouns(s: str | None) -> str | None:
        if s is None:
            return None
        # Remove title-cased words like "London" or "British" that tend to drift
        # Pexels results away from the current chosen city.
        out = re.sub(r"\b[A-Z][a-z]+\b", "", str(s))
        out = re.sub(r"\s+", " ", out).strip()
        return out or None

    def _strip_terms(kw: str | None, terms: list[str]) -> str | None:
        if kw is None:
            return None
        out = str(kw)
        for t in terms:
            t = (t or "").strip()
            if not t:
                continue
            # Remove occurrences (case-insensitive). Keep it simple: substring replace.
            out = re.sub(re.escape(t), "", out, flags=re.IGNORECASE)
        # Also remove duplicated commas/spaces that can be left behind.
        out = out.replace(",,", ",")
        out = re.sub(r"\s+", " ", out)
        out = re.sub(r"\s*,\s*", ", ", out)
        out = re.sub(r"\s+", " ", out).strip()
        return out or None

    def _strip_transport_words(kw: str | None) -> str | None:
        if kw is None:
            return None
        out = str(kw)
        transport_terms = [
            "flight",
            "flying",
            "airplane",
            "plane",
            "airport",
            "window",
            "cloud",
            "clouds",
            "cloudy",
            "sky",
            "skies",
            "takeoff",
            "take-off",
            "landing",
            "luggage",
            "boarding",
            "journey",
        ]
        for t in transport_terms:
            out = re.sub(rf"\b{re.escape(t)}\b", "", out, flags=re.IGNORECASE)
        out = re.sub(r"\s+", " ", out).strip(" ,")
        return out or None

    excel = Path(excel_path or config.TRAVEL_PRICES_EXCEL_PATH or "").expanduser()
    if not excel or not excel.is_file():
        raise RuntimeError(
            "Excel path missing. Pass excel_path to the tool OR set TRAVEL_PRICES_EXCEL_PATH in .env."
        )

    out_dir = Path(output_dir or config.MCP_REELS_DIR)
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    reel_name = f"mcp_reel_{_slugify(destination)}_{stamp}.mp4"
    out_mp4 = out_dir / reel_name

    rows = load_price_rows(excel)
    chosen = select_price_rows_for_destination(rows, destination)
    if target_total_segments is not None:
        k = max(1, min(int(target_total_segments), len(chosen)))
        chosen = chosen[:k]

    # If Groq includes a specific city name in keywords (e.g. "London"),
    # it can bias Pexels away from other chosen destinations in the same reel.
    # Strip all chosen city/country terms from keyword strings.
    all_city_terms = [
        str(r.destination or "").strip() for r in chosen if str(r.destination or "").strip()
    ]
    all_country_terms = [
        str(r.country or "").strip() for r in chosen if str(r.country or "").strip()
    ]
    terms_to_strip_global = [destination] + all_city_terms + all_country_terms

    image_keywords = _strip_transport_words(_strip_proper_nouns(image_keywords))
    video_keywords = _strip_transport_words(_strip_proper_nouns(video_keywords))

    # Decide segment types: aim for 2-3 videos total.
    n = len(chosen)
    target_video = random.randint(2, 3) if n >= 4 else max(1, n // 2)

    total_seconds = random.uniform(14.0, 24.0)
    # Slightly scale by segment count, but keep it non-fixed.
    total_seconds = max(12.0, min(30.0, total_seconds * (n / 5.0)))
    fps = 30
    seg_actual = _align_seconds_to_fps(total_seconds / float(n), fps=fps)
    xfade_dur = media_processor._reel_pick_xfade_seconds(seg_actual)  # type: ignore[attr-defined]
    out_duration = seg_actual * n - (n - 1) * xfade_dur

    work = out_dir / f"_work_{reel_name.replace('.mp4','')}"
    work.mkdir(parents=True, exist_ok=True)

    used_img_urls: set[str] = set()
    used_vid_urls: set[str] = set()

    segments: list[dict[str, Any]] = []
    client = httpx.Client(timeout=90.0, follow_redirects=True)
    try:
        # Fetch media for each chosen row.
        for i, r in enumerate(chosen):
            row_dest = str(r.destination or "").strip()
            row_country = str(r.country or "").strip()
            media_destination = row_dest
            if row_country and row_country.lower() not in media_destination.lower():
                media_destination = f"{row_dest}, {row_country}".strip(", ")

            row_img_keywords = _strip_terms(image_keywords, terms=terms_to_strip_global)
            row_vid_keywords = _strip_terms(video_keywords, terms=terms_to_strip_global)

            q_price = r.price_text or (f"{int(r.price)} €" if r.price is not None else None)
            title_price = q_price
            if title_price is None and r.price_range:
                title_price = str(r.price_range)

            # Render per-segment overlay card (text only). We will reuse for both image/video segments.
            overlay_png = work / f"overlay_{i:02d}.png"
            _render_segment_overlay_png(
                overlay_png,
                # Headline per segment: city + country when available.
                destination=_segment_fly_label(r, destination),
                price_text=title_price,
                origin=r.origin or origin,
            )

            entry_media = fetch_media_for_destination_entry(
                destination=media_destination,
                origin=r.origin or origin,
                price_text=title_price,
                mode=mode,
                image_keywords=row_img_keywords,
                video_keywords=row_vid_keywords,
                work_dir=work / f"media_{i:02d}",
                used_image_urls=used_img_urls,
                used_video_urls=used_vid_urls,
                client=client,
            )

            segments.append(
                {
                    "row": r,
                    "overlay_png": overlay_png,
                    "image_path": entry_media["image_path"],
                    "video_path": entry_media["video_path"],
                    "pexels_query_image": entry_media.get("pexels_query_image"),
                    "pexels_query_video": entry_media.get("pexels_query_video"),
                }
            )
    finally:
        client.close()

    # Pick which segments are videos.
    video_candidates = [s for s in segments if s["video_path"]]
    random.shuffle(video_candidates)
    video_selected = set(id(s) for s in video_candidates[:target_video])
    for s in segments:
        s["type"] = "video" if id(s) in video_selected else "image"

    # Build normalized segment MP4s.
    segment_paths: list[Path] = []
    for i, s in enumerate(segments):
        seg_path = work / f"seg_{i:02d}.mp4"
        overlay_png = s["overlay_png"]
        if s["type"] == "video" and s.get("video_path"):
            _normalize_and_encode_video_segment(
                Path(s["video_path"]),
                overlay_png,
                seg_path,
                seconds_each=seg_actual,
                context=f"mcp reel video segment {i + 1}/{n}",
            )
        else:
            _encode_image_segment(
                Path(s["image_path"]),
                overlay_png,
                seg_path,
                seconds_each=seg_actual,
                context=f"mcp reel image segment {i + 1}/{n}",
            )
        segment_paths.append(seg_path)

    # Concatenate segments with xfade.
    tmp_out = work / "reel_noaudio_concat.mp4"
    media_processor._xfade_concat_reel_segments(  # type: ignore[attr-defined]
        segment_paths,
        seg_actual,
        xfade_dur,
        out_duration,
        tmp_out,
        context="mcp travel reel xfade concat",
        transition_style="slideleft",
    )

    # Optional music (match the same semantics as `config.resolve_reel_music` used elsewhere).
    do_music = include_music
    if do_music is None:
        do_music = True
    if any(w in prompt.lower() for w in ("no music", "silent", "mute")):
        do_music = False

    if do_music:
        resolved = config.resolve_reel_music(music_path)
        if resolved is not None:
            media_processor._mux_music(  # type: ignore[attr-defined]
                tmp_out,
                resolved,
                out_mp4,
            )
        else:
            tmp_out.replace(out_mp4)
    else:
        tmp_out.replace(out_mp4)

    return {
        "output_path": str(out_mp4.resolve()),
        "destination": destination,
        "origin": origin,
        "segments": [
            {
                "destination": str(s["row"].destination),
                "origin": str(s["row"].origin),
                "price_text": s["row"].price_text or s["row"].price_range or str(s["row"].price),
                "type": s["type"],
                "pexels_query_image": s.get("pexels_query_image"),
                "pexels_query_video": s.get("pexels_query_video"),
            }
            for s in segments
        ],
        "timeline_seconds": {
            "segment_seconds": seg_actual,
            "xfade_seconds": xfade_dur,
            "total_seconds": out_duration,
        },
    }


def _align_seconds_to_fps(seconds: float, *, fps: int = 30) -> float:
    seconds = float(seconds)
    return max(1.0 / float(fps), int(round(seconds * fps)) / float(fps))

