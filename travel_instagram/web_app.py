"""
Small web UI for the travel Instagram generator.

Run from project root (recommended — avoids wrong ``travel_instagram`` package)::
  python -m uvicorn velo_web:app --reload --host 127.0.0.1 --port 8000

Or::
  uvicorn travel_instagram.web_app:app --reload --host 127.0.0.1 --port 8000
  (set PYTHONPATH to the repo root first)
"""

from __future__ import annotations

import asyncio
import hashlib
import io
import json
import logging
import re
import secrets
import shutil
import zipfile
from urllib.parse import unquote
from contextlib import asynccontextmanager
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Literal

import httpx
from fastapi import FastAPI, File, Form, HTTPException, Query, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from pydantic import BaseModel, Field
from starlette.templating import Jinja2Templates

from app.models.place import TravelMediaRequest, TravelMediaResponse
from app.services.aggregator import aggregate_travel_media

from travel_instagram import config
from travel_instagram import pipeline
from travel_instagram import groq_service
from travel_instagram import manual_reel_builder
from travel_instagram import manual_reel_autofill
from travel_instagram import media_processor
from travel_instagram import mcp_reel_tool
from travel_instagram import reels_catalog
from travel_instagram.instagram_post_export import safe_carousel_run_dir
from travel_instagram.instapost.router import router as instapost_router

logger = logging.getLogger(__name__)

# Reels AD / items_json: remote URL downloads per manual-reel build
_AD_REEL_MAX_URL_ITEMS = 28

TEMPLATES_DIR = Path(__file__).resolve().parent / "templates"
templates = Jinja2Templates(directory=str(TEMPLATES_DIR))


@asynccontextmanager
async def _velo_lifespan(app: FastAPI):
    paths = {getattr(r, "path", None) for r in app.routes if getattr(r, "path", None)}
    logger.warning(
        "Velo UI: %s | GET /ad-reels registered=%s",
        Path(__file__).resolve(),
        "/ad-reels" in paths,
    )
    yield


app = FastAPI(
    title="Travel Instagram Generator",
    description="Generate carousels and reels from a theme (Groq + Pexels + FFmpeg).",
    version="1.0.0",
    lifespan=_velo_lifespan,
)

config.ensure_output_dirs()
app.mount(
    "/media",
    StaticFiles(directory=str(config.OUTPUT_DIR)),
    name="media",
)

app.mount(
    "/music",
    StaticFiles(directory=str(config.MUSIC_LIBRARY_DIR)),
    name="music",
)

app.include_router(instapost_router)


def _abs_path_under_output(p: str | Path) -> Path | None:
    try:
        resolved = Path(p).resolve()
        out = config.OUTPUT_DIR.resolve()
        resolved.relative_to(out)
        return resolved
    except (ValueError, OSError):
        return None


def _to_media_url(abs_path: str | Path) -> str | None:
    p = _abs_path_under_output(abs_path)
    if p is None:
        return None
    rel = p.relative_to(config.OUTPUT_DIR.resolve())
    return "/media/" + rel.as_posix()


def _enrich_summary_for_web(summary: dict[str, Any]) -> dict[str, Any]:
    out = dict(summary)
    outputs = dict(summary.get("outputs") or {})

    slides = outputs.get("carousel_slides") or []
    outputs["carousel_slides_urls"] = [_to_media_url(s) for s in slides]

    reel = outputs.get("reel_video")
    outputs["reel_video_url"] = _to_media_url(reel) if reel else None

    sj = outputs.get("summary_json")
    outputs["summary_json_url"] = _to_media_url(sj) if sj else None

    out["outputs"] = outputs
    return out


class GenerateBody(BaseModel):
    theme: str = Field(
        ...,
        min_length=1,
        max_length=2000,
        description="Short theme or detailed media brief (place + requested shots); passed to Groq for Pexels queries.",
    )
    music_track_id: str | None = Field(
        default=None,
        max_length=512,
        description="Relative path under music/, __none__ for silence, null/__auto__ for .env/first file.",
    )


class McpReelBody(BaseModel):
    prompt: str = Field(..., min_length=1, max_length=800)
    music_track_id: str | None = Field(
        default=None,
        max_length=512,
        description="Relative path under music/ for audio, __none__ for silence, or null/__auto__ for automatic.",
    )


class AutofillReelMediaBody(BaseModel):
    theme: str = Field(
        ...,
        min_length=1,
        max_length=2000,
        description="Short theme or detailed media brief; Groq expands into per-scene Pexels search queries.",
    )
    max_items: int = Field(
        default=8,
        ge=1,
        le=20,
        description="Upper bound on downloaded clips; server bumps this when the theme includes 'top N places'.",
    )
    include_video: bool = Field(default=True, description="Download portrait videos too.")


class AdReelsZipItem(BaseModel):
    url: str = Field(..., min_length=8, max_length=4000)
    filename: str = Field(..., min_length=1, max_length=512)


class AdReelsZipBody(BaseModel):
    items: list[AdReelsZipItem] = Field(..., min_length=1, max_length=32)


class TravelBlogGenerateBody(BaseModel):
    """Groq → full HTML travel blog from title + parallel captions + image URLs."""

    title: str = Field(..., min_length=1, max_length=220)
    captions: list[str] = Field(default_factory=list, max_length=24)
    images: list[str] = Field(..., min_length=1, max_length=24)


class AdReelsLibrarySaveBody(BaseModel):
    """Persist prompt + travel-media snapshot: downloads clips into ``output/ad_reels_library/``."""

    model_config = {"extra": "ignore"}

    mode: Literal["all", "selected"] = "selected"
    selected_urls: list[str] = Field(default_factory=list, max_length=80)
    query: str = Field(default="", max_length=2000)
    tags: list[str] = Field(default_factory=list, max_length=30)
    orientation: str | None = Field(default=None, max_length=32)
    places: list[dict[str, Any]] = Field(default_factory=list)
    groq_places: list[dict[str, Any]] = Field(default_factory=list)
    search_plan: list[dict[str, Any]] = Field(default_factory=list)
    user_query: str = Field(default="", max_length=2000)
    groq_model: str | None = None
    pexels_calls_used: int = 0
    cache_hits: int = 0


class ExcelReelRowRequest(BaseModel):
    model_config = {"extra": "ignore"}
    destination: str = Field(..., min_length=1, max_length=2000)
    tags: str = Field(default="", max_length=500)
    orientation: str = Field(default="portrait", max_length=32)
    hook: str = Field(default="", max_length=500)
    pick: int = Field(default=2, ge=1, le=10)


_AD_REELS_LIB_MAX_FILES = 40
_SESSION_ID_SAFE = re.compile(r"^[a-zA-Z0-9._-]+$")


def _resolve_local_media_path(url: str) -> Path | None:
    """If ``url`` is a served ``/media/...`` path, return the file under ``OUTPUT_DIR``."""
    u = (url or "").strip()
    if not u.startswith("/media/"):
        return None
    rel = unquote(u[len("/media/") :].lstrip("/"))
    if not rel or ".." in rel.split("/"):
        return None
    p = (config.OUTPUT_DIR / rel).resolve()
    try:
        p.relative_to(config.OUTPUT_DIR.resolve())
    except ValueError:
        return None
    return p if p.is_file() else None


def _normalize_remote_media_url(url: str) -> str:
    u = (url or "").strip()
    if not u:
        return ""
    return u.split("?", 1)[0].rstrip("/")


def _find_ad_reels_library_local_path(remote_url: str) -> Path | None:
    """
    If ``remote_url`` matches a clip saved under ``output/ad_reels_library/``, return that file.

    Used so "Build reel" copies from disk instead of hitting Pexels again when the user
    has already saved the session (or the same URL exists in an older session).
    """
    raw = (remote_url or "").strip()
    if not raw.startswith(("http://", "https://")):
        return None
    lib = config.AD_REELS_LIBRARY_DIR.resolve()
    if not lib.is_dir():
        return None

    target_norm = _normalize_remote_media_url(raw)

    subdirs = [p for p in lib.iterdir() if p.is_dir()]
    subdirs.sort(key=lambda p: p.name, reverse=True)
    for sess in subdirs[:120]:
        sj = sess / "session.json"
        if not sj.is_file():
            continue
        try:
            doc = json.loads(sj.read_text(encoding="utf-8"))
        except (OSError, UnicodeDecodeError, json.JSONDecodeError):
            continue
        for place in doc.get("places") or []:
            if not isinstance(place, dict):
                continue
            for m in place.get("media") or []:
                if not isinstance(m, dict):
                    continue
                ru = str(m.get("remote_url") or "").strip()
                if not ru:
                    continue
                if _normalize_remote_media_url(ru) != target_norm and ru.rstrip("/") != raw.rstrip("/"):
                    continue
                media_url = str(m.get("url") or "").strip()
                if not media_url.startswith("/media/"):
                    continue
                p = _resolve_local_media_path(media_url)
                if p is not None and p.is_file():
                    return p

    # Library filenames embed sha256(remote_url)[:10] from save time.
    digest_candidates = {target_norm, raw}
    hits: list[Path] = []
    for cand in digest_candidates:
        if not cand:
            continue
        d10 = hashlib.sha256(cand.encode("utf-8")).hexdigest()[:10]
        for ext in (".jpg", ".jpeg", ".png", ".webp", ".mp4", ".webm", ".mov", ".m4v"):
            for p in lib.rglob(f"*_{d10}{ext}"):
                if p.is_file():
                    hits.append(p)
    if hits:
        return max(hits, key=lambda p: p.stat().st_mtime_ns)
    return None


def _slug_fs_segment(name: str, max_len: int = 44) -> str:
    t = re.sub(r"[^\w\-]+", "_", (name or "place").strip())
    return (t or "place")[:max_len]


def _zip_ad_reels_items_sync(items: list[tuple[str, str]]) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        with httpx.Client(timeout=120.0, follow_redirects=True) as client:
            for i, (url, filename) in enumerate(items):
                safe = Path(filename).name or f"file_{i}"
                arc = f"{i:02d}_{safe}"
                local = _resolve_local_media_path(url)
                if local is not None:
                    zf.write(local, arcname=arc)
                else:
                    r = client.get(url)
                    r.raise_for_status()
                    zf.writestr(arc, r.content)
    return buf.getvalue()


async def _persist_ad_reels_library(body: AdReelsLibrarySaveBody) -> dict[str, Any]:
    config.ensure_output_dirs()
    lib_root = config.AD_REELS_LIBRARY_DIR.resolve()
    lib_root.mkdir(parents=True, exist_ok=True)
    session_id = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ") + "_" + secrets.token_hex(4)
    session_dir = (lib_root / session_id).resolve()
    try:
        session_dir.relative_to(lib_root)
    except ValueError as e:
        raise HTTPException(status_code=400, detail="Invalid library path.") from e
    session_dir.mkdir(parents=True, exist_ok=True)

    selected_set = {u.strip() for u in body.selected_urls if u.strip()}
    places_out: list[dict[str, Any]] = []
    n_saved = 0

    for pi, place in enumerate(body.places):
        if not isinstance(place, dict):
            continue
        if n_saved >= _AD_REELS_LIB_MAX_FILES:
            break
        pname = str(place.get("name") or f"place_{pi}")
        pslug = _slug_fs_segment(pname) + f"_{pi:02d}"
        media_in = list(place.get("media") or [])
        media_out: list[dict[str, Any]] = []

        for mi, m in enumerate(media_in):
            if n_saved >= _AD_REELS_LIB_MAX_FILES:
                break
            if not isinstance(m, dict):
                continue
            url = str(m.get("url") or "").strip()
            if not url:
                continue
            if body.mode == "selected" and url not in selected_set:
                continue

            mtype = str(m.get("type") or "image").lower()
            ext = ".mp4" if mtype == "video" else ".jpg"
            low = url.split("?", 1)[0].lower()
            for cand in (".mp4", ".webm", ".mov", ".m4v", ".jpg", ".jpeg", ".png", ".webp"):
                if low.endswith(cand):
                    ext = cand
                    break
            digest = hashlib.sha256(url.encode("utf-8")).hexdigest()[:10]
            fname = f"{mtype}_{mi:02d}_{digest}{ext}"
            rel = Path("ad_reels_library") / session_id / pslug / fname
            dest = (config.OUTPUT_DIR / rel).resolve()
            try:
                dest.relative_to(config.OUTPUT_DIR.resolve())
            except ValueError as e:
                raise HTTPException(status_code=400, detail="Invalid media path.") from e
            dest.parent.mkdir(parents=True, exist_ok=True)
            local_src = _resolve_local_media_path(url)
            if local_src is not None:
                await asyncio.to_thread(shutil.copy2, local_src, dest)
            else:
                await asyncio.to_thread(media_processor.download_binary, url, dest)
            if not dest.is_file() or dest.stat().st_size == 0:
                continue
            n_saved += 1
            entry = dict(m)
            entry["url"] = "/media/" + rel.as_posix()
            entry["remote_url"] = url
            media_out.append(entry)

        if media_out:
            po = dict(place)
            po["media"] = media_out
            places_out.append(po)

    if not places_out:
        try:
            session_dir.rmdir()
        except OSError:
            pass
        raise HTTPException(
            status_code=400,
            detail="Nothing was saved. For “selected”, tick clips and try again; for “all”, ensure media is loaded.",
        )

    doc: dict[str, Any] = {
        "schema_version": 1,
        "session_id": session_id,
        "saved_at": datetime.now(timezone.utc).isoformat(),
        "request": {
            "query": body.query,
            "tags": body.tags,
            "orientation": body.orientation,
        },
        "user_query": body.user_query or body.query,
        "places": places_out,
        "groq_places": body.groq_places,
        "search_plan": body.search_plan,
        "groq_model": body.groq_model,
        "pexels_calls_used": body.pexels_calls_used,
        "cache_hits": body.cache_hits,
    }
    (session_dir / "session.json").write_text(json.dumps(doc, indent=2), encoding="utf-8")
    return {
        "session_id": session_id,
        "open_url": f"/ad-reels?session={session_id}",
        "library_url": "/ad-reels/library",
    }


@app.get("/", response_class=HTMLResponse)
async def index(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(
        "index.html",
        {
            "request": request,
            "title": "Travel Instagram Generator",
            "nav_active": "create",
        },
    )


@app.get("/reels", response_class=HTMLResponse)
async def reels_page(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(
        "reels.html",
        {
            "request": request,
            "title": "Reel library",
            "nav_active": "reels",
        },
    )


@app.get("/mcp-reels", response_class=HTMLResponse)
async def mcp_reels_page(request: Request) -> HTMLResponse:
    """Isolated UI for price reels via MCP tool."""
    return templates.TemplateResponse(
        "mcp_reels.html",
        {
            "request": request,
            "title": "Price Reels",
            "nav_active": "mcp_reels",
        },
    )


@app.get("/ad-reels", response_class=HTMLResponse)
async def ad_reels_page(request: Request) -> HTMLResponse:
    """Groq structured places + parallel Pexels preview (same registration style as ``/mcp-reels``)."""
    return templates.TemplateResponse(
        "reels_ad.html",
        {
            "request": request,
            "title": "Reels AD — Travel media",
            "nav_active": "ad_reels",
        },
    )


@app.get("/upload-reel", response_class=HTMLResponse)
async def upload_reel_page(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(
        "upload_reel.html",
        {
            "request": request,
            "title": "Generate Reel from Media",
            "nav_active": "upload_reel",
        },
    )


@app.get("/travel-blog", response_class=HTMLResponse)
async def travel_blog_page(request: Request) -> HTMLResponse:
    """Viral travel blog generator (Groq) with HTML preview and copy."""
    return templates.TemplateResponse(
        "travel_blog.html",
        {
            "request": request,
            "title": "Travel blog generator",
            "nav_active": "travel_blog",
        },
    )


@app.post("/api/blog/generate")
async def api_blog_generate(body: TravelBlogGenerateBody) -> JSONResponse:
    """Generate a full HTML blog page from title, image URLs, and captions (Groq)."""
    try:
        res = await asyncio.to_thread(
            groq_service.generate_travel_blog_html,
            body.title.strip(),
            list(body.images),
            list(body.captions),
        )
        return JSONResponse(content=res)
    except RuntimeError as e:
        raise HTTPException(status_code=502, detail=str(e)) from e
    except Exception:
        logger.exception("Blog generation failed")
        raise HTTPException(status_code=500, detail="Blog generation failed.") from None


@app.post("/api/ad-reels/travel-media", response_model=TravelMediaResponse)
async def api_ad_reels_travel_media(body: TravelMediaRequest) -> TravelMediaResponse:
    """
    Run Groq → 5 structured places → up to 20 parallel Pexels searches.
    Response includes groq_places, search_plan, and per-place media URLs.
    """
    q = (body.query or "").strip()
    if not q:
        raise HTTPException(status_code=400, detail="query is required.")
    try:
        async with httpx.AsyncClient(timeout=120.0, follow_redirects=True) as client:
            return await aggregate_travel_media(
                q,
                client,
                extra_tags=body.tags,
                orientation=body.orientation,
                download=body.download,
            )
    except RuntimeError as e:
        logger.warning("Reels AD travel media failed: %s", e)
        raise HTTPException(status_code=502, detail=str(e)) from e
    except httpx.HTTPError as e:
        logger.exception("Reels AD upstream HTTP error")
        raise HTTPException(status_code=502, detail=f"Upstream error: {e}") from e


@app.get("/ad-reels/library", response_class=HTMLResponse)
async def ad_reels_library_page(request: Request) -> HTMLResponse:
    """List saved Reels AD sessions (local media + stored prompt)."""
    return templates.TemplateResponse(
        "ad_reels_library.html",
        {
            "request": request,
            "title": "Reels AD — Saved sessions",
            "nav_active": "ad_reels_library",
        },
    )


def _resolve_queue_file() -> Path:
    raw = getattr(config, "TRAVEL_PRICES_EXCEL_PATH", None)
    if raw:
        p = Path(raw)
        # If the path exists as-is (native run), use it directly
        if p.exists():
            return p
        # Inside Docker the host path won't resolve — use just the filename
        # relative to the container's output dir (volume-mounted at OUTPUT_DIR)
        return config.OUTPUT_DIR / p.name
    return config.OUTPUT_DIR / "reels-queue.xlsx"

_REELS_QUEUE_FILE = _resolve_queue_file()
_QUEUE_COL_ALIASES: dict[str, list[str]] = {
    "destination": ["destination", "place", "query", "prompt", "location", "topic"],
    "status": ["status", "state"],
    "tags": ["tags", "extra_tags", "keywords"],
    "orientation": ["orientation", "orient"],
    "hook": ["hook", "hook_caption", "opening"],
    "pick": ["pick", "pick_per_place", "images_per_place"],
    "video_url": ["video_url", "url", "reel_url", "output"],
    "processed_at": ["processed_at", "done_at", "completed_at"],
}


def _queue_col_map(header: list[str]) -> dict[str, int]:
    idx: dict[str, int] = {}
    for key, aliases in _QUEUE_COL_ALIASES.items():
        for alias in aliases:
            if alias in header:
                idx[key] = header.index(alias)
                break
    return idx


@app.get("/api/excel-reels/next-pending")
async def excel_reels_next_pending() -> JSONResponse:
    """Return the first row in output/reels-queue.xlsx with status == 'pending'."""
    from openpyxl import load_workbook as _lw
    if not _REELS_QUEUE_FILE.exists():
        return JSONResponse({"pending": False, "message": "reels-queue.xlsx not found in output/"})
    try:
        wb = _lw(str(_REELS_QUEUE_FILE))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Cannot read reels-queue.xlsx: {exc}") from exc
    if not rows:
        return JSONResponse({"pending": False, "message": "Empty file"})

    header = [str(c or "").lower().strip() for c in rows[0]]
    col = _queue_col_map(header)
    if "destination" not in col:
        raise HTTPException(status_code=400, detail="reels-queue.xlsx missing 'destination' column in row 1.")
    if "status" not in col:
        raise HTTPException(status_code=400, detail="reels-queue.xlsx missing 'status' column in row 1.")

    def _cell(row: tuple, key: str, default: str = "") -> str:
        i = col.get(key)
        if i is None or i >= len(row):
            return default
        v = row[i]
        return str(v).strip() if v is not None else default

    for idx, row in enumerate(rows[1:]):
        if _cell(row, "status", "").lower() == "pending":
            pick_raw = _cell(row, "pick", "2")
            try:
                pick = max(1, min(10, int(float(pick_raw))))
            except (ValueError, TypeError):
                pick = 2
            orient = _cell(row, "orientation", "portrait").lower()
            if orient not in ("portrait", "landscape", "square"):
                orient = "portrait"
            return JSONResponse({
                "pending": True,
                "row_index": idx,
                "destination": _cell(row, "destination"),
                "tags": _cell(row, "tags"),
                "orientation": orient,
                "hook": _cell(row, "hook"),
                "pick": pick,
            })
    return JSONResponse({"pending": False, "message": "No pending rows"})


class MarkDoneBody(BaseModel):
    row_index: int = Field(..., ge=0)
    video_url: str = Field(default="")
    status: str = Field(default="done", max_length=32)


@app.post("/api/excel-reels/mark-done")
async def excel_reels_mark_done(body: MarkDoneBody) -> JSONResponse:
    """Update a row in output/reels-queue.xlsx: set status, video_url, processed_at."""
    from openpyxl import load_workbook as _lw
    if not _REELS_QUEUE_FILE.exists():
        raise HTTPException(status_code=404, detail="reels-queue.xlsx not found in output/")
    try:
        wb = _lw(str(_REELS_QUEUE_FILE))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Cannot read reels-queue.xlsx: {exc}") from exc
    if not rows:
        raise HTTPException(status_code=400, detail="Empty file")

    header = [str(c or "").lower().strip() for c in rows[0]]
    col = _queue_col_map(header)

    # Ensure columns exist, appending if needed
    def _ensure_col(key: str, label: str) -> int:
        if key in col:
            return col[key]
        new_idx = len(header)
        header.append(label)
        ws.cell(row=1, column=new_idx + 1, value=label)
        col[key] = new_idx
        return new_idx

    status_col = _ensure_col("status", "status")
    url_col = _ensure_col("video_url", "video_url")
    ts_col = _ensure_col("processed_at", "processed_at")

    xl_row = body.row_index + 2  # header=1, data starts at 2
    if xl_row > ws.max_row:
        raise HTTPException(status_code=400, detail=f"row_index {body.row_index} out of range")

    ws.cell(row=xl_row, column=status_col + 1, value=body.status)
    ws.cell(row=xl_row, column=url_col + 1, value=body.video_url)
    ws.cell(row=xl_row, column=ts_col + 1, value=datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"))
    try:
        wb.save(str(_REELS_QUEUE_FILE))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to save Excel: {exc}") from exc
    return JSONResponse({"ok": True, "row_index": body.row_index, "status": body.status})


# ── Instagram caption generation + preview ────────────────────────────────

_IG_PREVIEWS_DIR = config.OUTPUT_DIR / "instagram_previews"

_IG_CAPTION_SYSTEM = """\
You are a travel content writer creating short Instagram Reel captions that spark curiosity and drive website traffic.

The Reel already shows the destination visuals, location names, and travel highlights.
DO NOT recreate the itinerary. DO NOT list every attraction. DO NOT describe what is already in the video.

Your job is to spark curiosity, inspire the viewer, and send them to budgetwing.com to explore more and build their own itinerary.

Return ONLY valid JSON (no markdown, no code fences) with this exact structure:
{
  "title": "",
  "caption": "",
  "hashtags": ["", "", ""],
  "keywords": ""
}

=== TITLE (40-70 characters, max 12 words, max 1 emoji) ===
Curiosity-driven, destination-specific. Never generic. Examples:
- "Annecy is one of France's most underrated gems ✨"
- "This French town deserves a spot on your bucket list"
- "Planning a trip to France? Start here 📍"
- "You won't believe how beautiful this place is 👀"

=== CAPTION (60-100 words total — short, punchy, mobile-friendly) ===

4 parts, blank line between each:

PART 1 — Hook (1 line, under 15 words)
Scroll-stopping. Use a country flag or single emoji. Examples:
- "🇫🇷 This French town deserves a spot on your bucket list."
- "Don't skip this hidden gem on your next Europe trip."
- "Planning a trip to France? Start here. ✨"
- "You won't believe how beautiful this place is. 👀"

PART 2 — Short Description (2-3 sentences only)
What makes this destination special — atmosphere, what it feels like, why it's worth visiting.
Do NOT list places or attractions. Write to inspire, not inform.
Vary your language for every destination — never reuse the same sentences.

PART 3 — Website CTA (2-3 lines)
Direct the reader to budgetwing.com to explore more and build their own itinerary. Rotate between styles:
Option A:
"🌍 Want the complete guide?
✨ Discover more hidden gems, travel tips, and create your own personalized itinerary at budgetwing.com
🔗 Link in bio."
Option B:
"📍 Explore more places, travel smarter, and build your own itinerary at budgetwing.com
Link in bio."
Option C:
"✈️ Find cheap flights and plan your trip at budgetwing.com
🔗 Link in bio."
Adapt the wording naturally to the destination — don't copy-paste verbatim every time.

PART 4 — Engagement CTA (1 line only, rotate each time)
- "📌 Save this for your next adventure."
- "✈️ Share this with your travel partner."
- "❤️ Which destination should we feature next?"
- "👇 Would you visit [destination]?"

=== HASHTAGS (3-5 only, no # symbol) ===
Destination-specific and relevant. For Annecy: Annecy, FranceTravel, EuropeTrip, TravelGuide, BucketList
Avoid generic tags like "travel" or "wanderlust".

=== KEYWORDS ===
10-15 SEO keywords separated by "|", no # symbols.

RULES:
- Total caption must be 60-100 words. Count carefully.
- No itinerary. No bullet lists of places. No day-by-day structure.
- Every caption must feel fresh — vary hooks, descriptions, and CTAs across different destinations.
- Natural, conversational tone. Easy to read on mobile.
"""


class InstagramCaptionBody(BaseModel):
    destination: str = Field(..., min_length=1, max_length=500)
    places: list[str] = Field(default_factory=list)
    place_descriptions: list[dict] = Field(default_factory=list)
    video_url: str = Field(default="")


class InstagramPreviewSaveBody(BaseModel):
    destination: str = Field(..., min_length=1)
    video_url: str = Field(default="")
    title: str = Field(default="")
    caption: str = Field(..., min_length=1)
    formatted_post: str = Field(default="")
    hashtags: list[str] = Field(default_factory=list)
    keywords: str = Field(default="")
    row_index: int = Field(default=-1)
    reel_duration_secs: float = Field(default=0.0)
    # YouTube-optimized metadata (optional — generated by /api/youtube/generate-metadata)
    youtube_title: str = Field(default="")
    youtube_description: str = Field(default="")
    youtube_tags: list[str] = Field(default_factory=list)
    youtube_pinned_comment: str = Field(default="")
    youtube_title_variations: list[str] = Field(default_factory=list)


@app.post("/api/instagram/generate-caption")
async def instagram_generate_caption(body: InstagramCaptionBody) -> JSONResponse:
    """Call Groq to produce a humanized caption, hashtags, and alt text for an Instagram post."""
    from groq import Groq as _Groq

    key = config.GROQ_API_KEY
    if not key:
        raise HTTPException(status_code=500, detail="GROQ_API_KEY not configured.")

    # Collect destination vibe from place descriptions for atmosphere context
    vibes = [pd.get("vibe", "") for pd in body.place_descriptions if pd.get("vibe")]
    vibe_line = " | ".join(vibes[:3]) if vibes else ""

    user_msg = (
        f"Destination: {body.destination}\n"
        + (f"Atmosphere / vibe: {vibe_line}\n" if vibe_line else "")
        + "\nWrite a short Instagram Reel caption (60-100 words) that sparks curiosity about this destination "
        "and drives viewers to visit budgetwing.com to explore more and build their own itinerary. "
        "Do NOT list attractions or recreate the itinerary — the Reel already shows that. "
        "Focus on inspiring the viewer with atmosphere and emotion, then direct them to the website."
    )

    try:
        client = _Groq(api_key=key)
        completion = await asyncio.to_thread(
            client.chat.completions.create,
            model=config.GROQ_MODEL,
            messages=[
                {"role": "system", "content": _IG_CAPTION_SYSTEM},
                {"role": "user", "content": user_msg},
            ],
            temperature=0.9,
            max_tokens=1024,
            response_format={"type": "json_object"},
        )
        raw = completion.choices[0].message.content or ""
        data = json.loads(raw)
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"Groq caption generation failed: {exc}") from exc

    title = str(data.get("title") or "").strip()
    caption = str(data.get("caption") or "").strip()
    hashtags = [str(h).strip().lstrip("#") for h in (data.get("hashtags") or []) if str(h).strip()][:5]
    keywords = str(data.get("keywords") or "").strip()

    hashtag_block = " ".join(f"#{h}" for h in hashtags)
    formatted_post = f"{caption}\n\n.\n.\n.\n{hashtag_block}"
    char_count = len(formatted_post)

    return JSONResponse({
        "title": title,
        "caption": caption,
        "hashtags": hashtags,
        "hashtag_block": hashtag_block,
        "keywords": keywords,
        "formatted_post": formatted_post,
        "char_count": char_count,
        "within_limit": char_count <= 2200,
        "groq_model": config.GROQ_MODEL,
    })


@app.post("/api/instagram/preview/save")
async def instagram_preview_save(body: InstagramPreviewSaveBody) -> JSONResponse:
    """Persist a preview record to disk and return its preview_id + URL."""
    _IG_PREVIEWS_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    slug = re.sub(r"[^a-z0-9]+", "-", body.destination.lower())[:40].strip("-")
    preview_id = f"{ts}_{slug}"
    preview_file = _IG_PREVIEWS_DIR / f"{preview_id}.json"

    hashtag_block = " ".join(f"#{h}" for h in body.hashtags)
    formatted_post = body.formatted_post or f"{body.caption}\n\n.\n.\n.\n{hashtag_block}"

    record = {
        "preview_id": preview_id,
        "destination": body.destination,
        "video_url": body.video_url,
        "title": body.title,
        "caption": body.caption,
        "hashtags": body.hashtags,
        "hashtag_block": hashtag_block,
        "keywords": body.keywords,
        "formatted_post": formatted_post,
        "char_count": len(formatted_post),
        "row_index": body.row_index,
        "reel_duration_secs": body.reel_duration_secs,
        "status": "pending_review",
        "created_at": datetime.now(timezone.utc).isoformat(),
        # YouTube-optimized metadata (empty strings if not generated)
        "youtube_title": body.youtube_title,
        "youtube_description": body.youtube_description,
        "youtube_tags": body.youtube_tags,
        "youtube_pinned_comment": body.youtube_pinned_comment,
        "youtube_title_variations": body.youtube_title_variations,
    }

    preview_file.write_text(json.dumps(record, indent=2, ensure_ascii=False), encoding="utf-8")
    return JSONResponse({
        "ok": True,
        "preview_id": preview_id,
        "preview_url": f"/instagram-preview/{preview_id}",
    })


@app.get("/api/instagram/preview/{preview_id}")
async def instagram_preview_get(preview_id: str) -> JSONResponse:
    if not re.fullmatch(r"[A-Za-z0-9_\-]+", preview_id):
        raise HTTPException(status_code=400, detail="Invalid preview_id.")
    f = _IG_PREVIEWS_DIR / f"{preview_id}.json"
    if not f.exists():
        raise HTTPException(status_code=404, detail="Preview not found.")
    return JSONResponse(json.loads(f.read_text(encoding="utf-8")))


@app.get("/api/instagram/credentials-status")
async def instagram_credentials_status() -> JSONResponse:
    from travel_instagram import instagram_service
    configured = instagram_service.instagram_credentials_configured()
    has_base_url = bool(config.PUBLIC_APP_BASE_URL)
    return JSONResponse({
        "configured": configured,
        "has_public_url": has_base_url,
        "ready": configured and has_base_url,
    })


@app.post("/api/instagram/publish/{preview_id}")
async def instagram_publish(preview_id: str) -> JSONResponse:
    """Publish an approved preview to Instagram Reels via the Graph API."""
    from travel_instagram import instagram_service

    if not re.fullmatch(r"[A-Za-z0-9_\-]+", preview_id):
        raise HTTPException(status_code=400, detail="Invalid preview_id.")

    f = _IG_PREVIEWS_DIR / f"{preview_id}.json"
    if not f.exists():
        raise HTTPException(status_code=404, detail="Preview not found.")

    preview = json.loads(f.read_text(encoding="utf-8"))

    if preview.get("status") == "published":
        return JSONResponse({
            "ok": True,
            "already_published": True,
            "ig_media_id": preview.get("ig_media_id"),
        })

    if not instagram_service.instagram_credentials_configured():
        raise HTTPException(
            status_code=503,
            detail="Instagram not configured. Set IG_USER_ID and IG_ACCESS_TOKEN in .env.",
        )

    base_url = config.PUBLIC_APP_BASE_URL
    if not base_url:
        raise HTTPException(
            status_code=503,
            detail=(
                "PUBLIC_APP_BASE_URL is not set. Meta needs a public HTTPS URL to fetch the video. "
                "Run ngrok (ngrok http 8000) and set PUBLIC_APP_BASE_URL=https://your-id.ngrok.io in .env."
            ),
        )

    video_url = (preview.get("video_url") or "").strip()
    if not video_url:
        raise HTTPException(status_code=400, detail="Preview has no video_url.")

    public_video_url = base_url + video_url if video_url.startswith("/") else video_url
    if not public_video_url.startswith("https://"):
        raise HTTPException(
            status_code=400,
            detail=f"Video URL must be HTTPS. Got: {public_video_url}",
        )

    caption_text = preview.get("formatted_post") or preview.get("caption") or ""

    try:
        result = await asyncio.to_thread(
            instagram_service.publish_reel,
            video_url=public_video_url,
            caption=caption_text,
        )
    except (RuntimeError, ValueError) as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc

    ig_media_id = result.get("id", "")
    preview["status"] = "published"
    preview["ig_media_id"] = ig_media_id
    preview["published_at"] = datetime.now(timezone.utc).isoformat()
    f.write_text(json.dumps(preview, indent=2, ensure_ascii=False), encoding="utf-8")

    return JSONResponse({"ok": True, "ig_media_id": ig_media_id, "published_at": preview["published_at"]})


_YT_METADATA_SYSTEM = """You are an expert YouTube Shorts growth strategist and SEO specialist \
who manages viral travel channels with millions of subscribers.

Your goal is to create metadata that feels authentic, drives clicks, and ranks in YouTube search. \
Write like a top travel creator — not a brand or travel agency. \
Use ONLY the information provided. Do not invent facts or locations.

Return ONLY valid JSON with these exact keys:
{
  "title": "",
  "description": "",
  "hashtags": [],
  "tags": [],
  "pinned_comment": "",
  "title_variations": ["", "", ""]
}

Rules:

Title (under 60 characters):
- Hook-driven and emotionally engaging — make viewers feel something
- Specific to the destination, not generic
- Use power words: hidden, secret, honest, stunning, unexpected, underrated
- Max 1 emoji if it adds impact

Description (150–300 characters):
- Written like a travel creator's personal note, not a press release
- Start with a hook sentence, then weave in location + keywords naturally
- End with a soft CTA: "Save this for your trip 📌" or "Drop a comment below 👇"
- Include the destination name and 1–2 landmark/experience keywords

Hashtags (5–7 tags, always include #Shorts and #Travel):
- Mix location-specific, niche travel, and broad reach tags

Tags (12–18 strings without #):
- Highly relevant SEO tags covering destination, experiences, travel style, season

Pinned comment:
- Conversational and warm — ask a question or share a personal tip
- Encourages replies: "Where would you go first?" / "Have you been here?"
- 1–2 sentences max

Title variations (exactly 3):
1. Curiosity-driven — makes viewer need to know more
2. Emotional — taps into wanderlust or FOMO
3. Search-friendly — matches what travellers actually type

Do not use misleading clickbait. Do not add information not in the input."""


class YouTubeMetadataBody(BaseModel):
    title: str = Field(default="")
    video_summary: str = Field(default="")
    location: str = Field(default="")
    landmarks: str = Field(default="")
    mood: str = Field(default="")
    keywords: str = Field(default="")


@app.post("/api/youtube/generate-metadata")
async def youtube_generate_metadata(body: YouTubeMetadataBody) -> JSONResponse:
    """Call Groq to generate SEO-optimised YouTube Shorts metadata."""
    from groq import Groq as _Groq

    key = config.GROQ_API_KEY
    if not key:
        raise HTTPException(status_code=500, detail="GROQ_API_KEY not configured.")

    user_msg = (
        f"Title: {body.title}\n"
        f"Video Summary: {body.video_summary}\n"
        f"Location: {body.location}\n"
        f"Landmarks: {body.landmarks}\n"
        f"Mood: {body.mood}\n"
        f"Keywords: {body.keywords}"
    )

    try:
        client = _Groq(api_key=key)
        completion = await asyncio.to_thread(
            client.chat.completions.create,
            model=config.GROQ_MODEL,
            messages=[
                {"role": "system", "content": _YT_METADATA_SYSTEM},
                {"role": "user", "content": user_msg},
            ],
            temperature=0.8,
            max_tokens=1024,
            response_format={"type": "json_object"},
        )
        raw = completion.choices[0].message.content or ""
        data = json.loads(raw)
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"Groq YouTube metadata generation failed: {exc}") from exc

    return JSONResponse({
        "youtube_title": str(data.get("title") or "").strip()[:100],
        "youtube_description": str(data.get("description") or "").strip()[:5000],
        "youtube_hashtags": [str(h).strip().lstrip("#") for h in (data.get("hashtags") or [])],
        "youtube_tags": [str(t).strip() for t in (data.get("tags") or [])],
        "youtube_pinned_comment": str(data.get("pinned_comment") or "").strip(),
        "youtube_title_variations": [str(v).strip() for v in (data.get("title_variations") or [])],
    })


@app.get("/api/youtube/channel")
async def youtube_channel_info() -> JSONResponse:
    """Return the YouTube channel connected to the configured credentials."""
    from travel_instagram import youtube_service
    if not youtube_service.youtube_credentials_configured():
        raise HTTPException(status_code=503, detail="YouTube credentials not configured.")
    try:
        info = await asyncio.to_thread(youtube_service.get_channel_info)
        return JSONResponse(info)
    except RuntimeError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc


@app.post("/api/youtube/publish/{preview_id}")
async def youtube_publish(preview_id: str) -> JSONResponse:
    """Publish a saved preview as a YouTube Short."""
    from travel_instagram import youtube_service

    if not re.fullmatch(r"[A-Za-z0-9_\-]+", preview_id):
        raise HTTPException(status_code=400, detail="Invalid preview_id.")

    f = _IG_PREVIEWS_DIR / f"{preview_id}.json"
    if not f.exists():
        raise HTTPException(status_code=404, detail="Preview not found.")

    preview = json.loads(f.read_text(encoding="utf-8"))

    if preview.get("youtube_video_id"):
        return JSONResponse({
            "ok": True,
            "already_published": True,
            "youtube_video_id": preview["youtube_video_id"],
            "youtube_url": preview.get("youtube_url", ""),
        })

    if not youtube_service.youtube_credentials_configured():
        raise HTTPException(
            status_code=503,
            detail="YouTube not configured. Set YOUTUBE_CLIENT_ID, YOUTUBE_CLIENT_SECRET, YOUTUBE_REFRESH_TOKEN in .env. Run scripts/youtube_auth.py to get the refresh token.",
        )

    video_url = (preview.get("video_url") or "").strip()
    if not video_url:
        raise HTTPException(status_code=400, detail="Preview has no video_url.")

    # Resolve /media/... URL to actual file path
    rel = video_url.lstrip("/")
    if rel.startswith("media/"):
        rel = rel[len("media/"):]
    video_path = config.OUTPUT_DIR / rel
    if not video_path.is_file():
        raise HTTPException(status_code=404, detail=f"Video file not found at {video_path}")

    # Prefer YouTube-optimized metadata when available, fall back to Instagram caption data
    title = preview.get("youtube_title") or preview.get("title") or preview.get("destination") or "Travel Reel"
    description = preview.get("youtube_description") or preview.get("formatted_post") or preview.get("caption") or ""
    hashtags = preview.get("youtube_tags") or preview.get("hashtags") or []

    try:
        result = await asyncio.to_thread(
            youtube_service.publish_short,
            video_path=video_path,
            title=title,
            description=description,
            tags=hashtags,
            privacy="public",
        )
    except (RuntimeError, ValueError) as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc

    preview["youtube_video_id"] = result.get("youtube_video_id", "")
    preview["youtube_url"] = result.get("youtube_url", "")
    preview["youtube_published_at"] = datetime.now(timezone.utc).isoformat()
    f.write_text(json.dumps(preview, indent=2, ensure_ascii=False), encoding="utf-8")

    return JSONResponse({
        "ok": True,
        "youtube_video_id": result.get("youtube_video_id", ""),
        "youtube_url": result.get("youtube_url", ""),
        "title": result.get("title", ""),
    })


@app.get("/instagram-preview/{preview_id}", response_class=HTMLResponse)
async def instagram_preview_page(preview_id: str, request: Request) -> HTMLResponse:
    if not re.fullmatch(r"[A-Za-z0-9_\-]+", preview_id):
        raise HTTPException(status_code=400, detail="Invalid preview_id.")
    f = _IG_PREVIEWS_DIR / f"{preview_id}.json"
    if not f.exists():
        raise HTTPException(status_code=404, detail="Preview not found.")
    data = json.loads(f.read_text(encoding="utf-8"))
    return templates.TemplateResponse(
        "instagram_preview.html",
        {"request": request, "preview": data, "nav_active": "instagram_preview"},
    )


@app.get("/excel-reels", response_class=HTMLResponse)
async def excel_reels_page(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(
        "excel_reels.html",
        {"request": request, "title": "Excel → Reels", "nav_active": "excel_reels"},
    )


@app.get("/api/excel-reels/template")
async def excel_reels_template() -> Response:
    """Return a sample .xlsx the user can fill in and upload."""
    import io as _io
    from openpyxl import Workbook as _Workbook
    wb = _Workbook()
    ws = wb.active
    ws.title = "Reels"
    ws.append(["destination", "tags", "orientation", "hook", "pick"])
    ws.append(["Top places to visit in Bali", "sunset, beach, aerial", "portrait", "Bali will steal your heart", 2])
    ws.append(["Hidden gems of Portugal", "drone, coastline, historic", "landscape", "Portugal you've never seen", 2])
    ws.append(["Best beaches in Thailand", "turquoise water, islands", "portrait", "Thailand awaits", 2])
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="velo-reels-template.xlsx"'},
    )


@app.post("/api/excel-reels/parse")
async def excel_reels_parse(file: UploadFile = File(...)) -> JSONResponse:
    """Parse an uploaded Excel file and return destination rows."""
    import io as _io
    from openpyxl import load_workbook as _lw
    data = await file.read()
    try:
        wb = _lw(_io.BytesIO(data))
        ws = wb.active
        rows_raw = list(ws.iter_rows(values_only=True))
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Cannot read Excel file: {exc}") from exc
    if not rows_raw:
        return JSONResponse({"rows": [], "count": 0})

    col_aliases: dict[str, list[str]] = {
        "destination": ["destination", "place", "query", "prompt", "location", "topic"],
        "tags": ["tags", "extra_tags", "keywords", "tag"],
        "orientation": ["orientation", "orient"],
        "hook": ["hook", "hook_caption", "opening", "intro"],
        "pick": ["pick", "pick_per_place", "images_per_place", "media_count"],
    }
    header = [str(c or "").lower().strip() for c in rows_raw[0]]
    col_idx: dict[str, int] = {}
    for key, aliases in col_aliases.items():
        for alias in aliases:
            if alias in header:
                col_idx[key] = header.index(alias)
                break

    def _get(raw: tuple, key: str, default: str = "") -> str:
        idx = col_idx.get(key)
        if idx is None or idx >= len(raw):
            return default
        v = raw[idx]
        return str(v).strip() if v is not None else default

    rows: list[dict] = []
    for raw in rows_raw[1:]:
        dest = _get(raw, "destination")
        if not dest:
            continue
        pick_raw = _get(raw, "pick", "2")
        try:
            pick = max(1, min(10, int(float(pick_raw))))
        except (ValueError, TypeError):
            pick = 2
        orient = _get(raw, "orientation", "portrait").lower()
        if orient not in ("portrait", "landscape", "square"):
            orient = "portrait"
        rows.append({
            "destination": dest,
            "tags": _get(raw, "tags"),
            "orientation": orient,
            "hook": _get(raw, "hook"),
            "pick": pick,
        })
    return JSONResponse({"rows": rows, "count": len(rows)})


@app.post("/api/excel-reels/run-one")
async def excel_reels_run_one(body: ExcelReelRowRequest) -> JSONResponse:
    """Full server-side pipeline for one Excel row: Groq → Pexels → pick N per place → build reel."""
    from app.services.aggregator import aggregate_travel_media as _agg
    tags = [t.strip() for t in body.tags.split(",") if t.strip()] if body.tags else []
    orientation = body.orientation if body.orientation in ("portrait", "landscape", "square") else None

    async with httpx.AsyncClient(timeout=120.0) as client:
        result = await _agg(
            user_query=body.destination,
            client=client,
            extra_tags=tags,
            orientation=orientation,
        )

    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    temp_dir = config.OUTPUT_DIR / "manual_reels" / f"excel_{ts}"
    temp_dir.mkdir(parents=True, exist_ok=True)

    media_paths: list[Path] = []
    titles: list[str] = []
    caption_texts: list[str] = []
    for place in result.places:
        for m in (place.media or [])[: body.pick]:
            ext = ".mp4" if m.type == "video" else ".jpg"
            digest = hashlib.sha256(m.url.encode()).hexdigest()[:12]
            dest = temp_dir / f"media_{len(media_paths):02d}_{digest}{ext}"
            try:
                await asyncio.to_thread(media_processor.download_binary, m.url, dest)
            except Exception:
                continue
            if dest.is_file() and dest.stat().st_size > 0:
                media_paths.append(dest)
                titles.append(place.name or "")
                caption_texts.append((place.caption_text or "").strip())

    if not media_paths:
        raise HTTPException(status_code=502, detail="No media could be downloaded for this destination.")

    overlay_positions = [(0.5, 0.15)] * len(media_paths)
    overlay_font_scales = [1.0] * len(media_paths)
    captions: list[str] = [""] * len(media_paths)

    try:
        res = await asyncio.to_thread(
            manual_reel_builder.build_manual_reel,
            uploads_dir=temp_dir,
            media_paths=media_paths,
            captions=captions,
            music_track_id=None,
            transition_type="slideleft",
            transition_speed="default",
            transition_xfade_scale=None,
            overlay_positions=overlay_positions,
            overlay_font_scales=overlay_font_scales,
            titles=titles,
            caption_texts=caption_texts,
            hook_caption=body.hook or "",
            hook_seconds=3.0,
            image_segment_seconds=3.0,
            video_segment_seconds=5.0,
            show_branding=True,
        )
        out = dict(res)
        out["video_url"] = _to_media_url(out.get("output_path") or "") or None
        return JSONResponse(content=out)
    except RuntimeError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc
    except Exception:
        logger.exception("Excel reel generation failed for %s", body.destination)
        raise HTTPException(status_code=500, detail="Reel generation failed.") from None


@app.post("/api/ad-reels/library/save")
async def api_ad_reels_library_save(body: AdReelsLibrarySaveBody) -> JSONResponse:
    """Download selected or all remote clips into ``output/ad_reels_library/<session>/`` and write ``session.json``."""
    if not body.places:
        raise HTTPException(status_code=400, detail="places is required (run Fetch media first).")
    try:
        meta = await _persist_ad_reels_library(body)
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("ad-reels library save failed")
        raise HTTPException(status_code=502, detail=str(e)) from e
    return JSONResponse(content=meta)


@app.get("/api/ad-reels/library")
async def api_ad_reels_library_list() -> JSONResponse:
    config.ensure_output_dirs()
    lib = config.AD_REELS_LIBRARY_DIR
    sessions: list[dict[str, Any]] = []
    if lib.is_dir():
        for d in sorted(lib.iterdir(), key=lambda x: x.name, reverse=True):
            if not d.is_dir():
                continue
            meta_path = d / "session.json"
            if not meta_path.is_file():
                continue
            try:
                doc = json.loads(meta_path.read_text(encoding="utf-8"))
            except (OSError, json.JSONDecodeError):
                continue
            req = doc.get("request") or {}
            sessions.append(
                {
                    "session_id": doc.get("session_id", d.name),
                    "saved_at": doc.get("saved_at", ""),
                    "query": req.get("query", ""),
                }
            )
    return JSONResponse(content={"sessions": sessions})


@app.get("/api/ad-reels/library/{session_id}")
async def api_ad_reels_library_get(session_id: str) -> JSONResponse:
    if not _SESSION_ID_SAFE.match(session_id):
        raise HTTPException(status_code=400, detail="Invalid session id.")
    lib_root = config.AD_REELS_LIBRARY_DIR.resolve()
    p = (lib_root / session_id / "session.json").resolve()
    try:
        p.relative_to(lib_root)
    except ValueError as e:
        raise HTTPException(status_code=400, detail="Invalid session path.") from e
    if not p.is_file():
        raise HTTPException(status_code=404, detail="Session not found.")
    try:
        doc = json.loads(p.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as e:
        raise HTTPException(status_code=500, detail="Corrupt session file.") from e
    return JSONResponse(content=doc)


@app.post("/api/ad-reels/library/zip")
async def api_ad_reels_library_zip(body: AdReelsZipBody) -> Response:
    """ZIP remote Pexels URLs (e.g. all or selected tiles) for one browser download."""
    pairs = [(it.url.strip(), it.filename.strip()) for it in body.items if it.url.strip()]
    if not pairs:
        raise HTTPException(status_code=400, detail="No URLs to zip.")
    try:
        raw = await asyncio.to_thread(_zip_ad_reels_items_sync, pairs)
    except httpx.HTTPError as e:
        raise HTTPException(status_code=502, detail=f"Download failed: {e}") from e
    except Exception as e:
        logger.exception("ad-reels zip failed")
        raise HTTPException(status_code=502, detail=str(e)) from e
    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    return Response(
        content=raw,
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="ad-reels_{ts}.zip"',
        },
    )


@app.post("/api/mcp-reels/generate")
async def api_mcp_reels_generate(body: McpReelBody) -> JSONResponse:
    prompt = (body.prompt or "").strip()
    if not prompt:
        raise HTTPException(status_code=400, detail="Prompt is required.")

    music_track_id = body.music_track_id
    if music_track_id == "__auto__":
        music_track_id = None
    if isinstance(music_track_id, str) and music_track_id.strip() == "":
        music_track_id = None

    try:
        refined: dict[str, Any] | None = None
        try:
            refined = await asyncio.to_thread(groq_service.parse_reel_prompt, prompt)
        except Exception:
            refined = None
        refined_prompt = ((refined or {}).get("refined_prompt") or prompt).strip()
        search_destination = (refined or {}).get("destination")
        search_origin = (refined or {}).get("origin")
        search_mode = (refined or {}).get("mode")
        image_keywords = (refined or {}).get("image_keywords")
        video_keywords = (refined or {}).get("video_keywords")

        res = await asyncio.to_thread(
            mcp_reel_tool.generate_travel_reel_from_prompt,
            refined_prompt,
            music_path=music_track_id,
            search_destination=search_destination,
            search_origin=search_origin,
            search_mode=search_mode,
            image_keywords=image_keywords,
            video_keywords=video_keywords,
        )

        out = dict(res)
        out["video_url"] = _to_media_url(out.get("output_path") or "") or None
        out["refined"] = refined
        return JSONResponse(content=out)
    except RuntimeError as e:
        raise HTTPException(status_code=502, detail=str(e)) from e
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("MCP reel generation failed")
        raise HTTPException(status_code=500, detail=str(e)) from e


@app.post("/api/upload-reel/generate")
async def api_upload_reel_generate(
    files: list[UploadFile] = File(default_factory=list),
    captions_json: str = Form(default="[]"),
    items_json: str = Form(default=""),
    music_track_id: str | None = Form(default=None),
    transition_type: str = Form(default="slideleft"),
    transition_speed: str = Form(default="default"),
    transition_xfade_scale: str | None = Form(default=None),
    overlay_font_scale: str = Form(default="1.0"),
    hook_caption: str = Form(default=""),
    hook_seconds: str = Form(default="3"),
    clip_seconds_image: str = Form(default="3"),
    clip_seconds_video: str = Form(default="5"),
    overlay_anchor_x: str = Form(default="0.5"),
    overlay_anchor_y: str = Form(default="0.15"),
    show_branding: str = Form(default="1"),
    music_volume: str = Form(default="0.3"),
) -> JSONResponse:
    if music_track_id == "__auto__":
        music_track_id = None
    if isinstance(music_track_id, str) and music_track_id.strip() == "":
        music_track_id = None

    base_autofill = config.OUTPUT_DIR / "manual_reels" / "autofill"

    allowed = {".jpg", ".jpeg", ".png", ".webp", ".mp4", ".mov", ".m4v", ".webm"}
    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    temp_dir = config.OUTPUT_DIR / "manual_reels" / f"tmp_{ts}"
    temp_dir.mkdir(parents=True, exist_ok=True)

    uploaded_paths: list[Path] = []
    for i, up in enumerate(files):
        suffix = Path(up.filename or "").suffix.lower()
        if suffix not in allowed:
            raise HTTPException(
                status_code=400,
                detail=f"Unsupported file type '{suffix}' for '{up.filename}'.",
            )
        name = f"upload_{i:02d}_{Path(up.filename or 'asset').stem}{suffix}"
        out = temp_dir / name
        data = await up.read()
        out.write_bytes(data)
        uploaded_paths.append(out)

    media_paths: list[Path] = []
    captions: list[str] = []
    titles: list[str] = []
    caption_texts: list[str] = []
    overlay_positions: list[tuple[float, float]] = []
    overlay_font_scales: list[float] = []
    try:
        requested_font_scale = float((overlay_font_scale or "1.0").strip() or "1.0")
    except ValueError:
        requested_font_scale = 1.0
    requested_font_scale = max(0.6, min(1.7, requested_font_scale))

    xfade_scale_opt: float | None = None
    if transition_xfade_scale is not None and str(transition_xfade_scale).strip() != "":
        try:
            xfade_scale_opt = float(str(transition_xfade_scale).strip())
        except ValueError:
            xfade_scale_opt = None
    if xfade_scale_opt is not None:
        xfade_scale_opt = max(0.45, min(1.65, xfade_scale_opt))

    hook_sec_f = 3.0
    try:
        hook_sec_f = float((hook_seconds or "3").strip() or "3")
    except ValueError:
        hook_sec_f = 3.0
    hook_sec_f = max(0.5, min(12.0, hook_sec_f))

    try:
        clip_img = float((clip_seconds_image or "3").strip() or "3")
    except ValueError:
        clip_img = 3.0
    try:
        clip_vid = float((clip_seconds_video or "5").strip() or "5")
    except ValueError:
        clip_vid = 5.0
    clip_img = max(0.5, min(90.0, clip_img))
    clip_vid = max(0.5, min(90.0, clip_vid))

    try:
        oax = float((overlay_anchor_x or "0.5").strip() or "0.5")
    except ValueError:
        oax = 0.5
    try:
        oay = float((overlay_anchor_y or "0.15").strip() or "0.15")
    except ValueError:
        oay = 0.15
    oax = max(0.05, min(0.95, oax))
    oay = max(0.05, min(0.92, oay))

    _sb = str(show_branding or "1").strip().lower()
    show_brand_on_reel = _sb not in ("0", "false", "no", "off")

    if items_json:
        try:
            parsed_items = json.loads(items_json)
        except json.JSONDecodeError as e:
            raise HTTPException(status_code=400, detail="Invalid items_json payload.") from e
        if not isinstance(parsed_items, list) or not parsed_items:
            raise HTTPException(status_code=400, detail="items_json must be a non-empty list.")

        _url_n = sum(
            1
            for x in parsed_items
            if isinstance(x, dict) and str(x.get("source") or "").strip().lower() == "url"
        )
        if _url_n > _AD_REEL_MAX_URL_ITEMS:
            raise HTTPException(
                status_code=400,
                detail=f"At most {_AD_REEL_MAX_URL_ITEMS} url items per reel (got {_url_n}).",
            )

        for it in parsed_items:
            if not isinstance(it, dict):
                raise HTTPException(status_code=400, detail="items_json entries must be objects.")
            src = str(it.get("source") or "").strip()
            caption = str(it.get("caption") or "")
            title = str(it.get("title") or "").strip()
            ct_raw = str(it.get("caption_text") or "").strip()
            if it.get("show_caption_text") is False:
                ct_raw = ""
            ct_raw = groq_service.normalize_reel_caption_text(ct_raw, max_words=15)
            if len(ct_raw) > 420:
                ct_raw = ct_raw[:420]
            overlay_x = it.get("overlay_x", 0.5)
            overlay_y = it.get("overlay_y", 0.15)
            font_scale = it.get("font_scale", 1.0)

            if src == "server":
                asset_id = str(it.get("asset_id") or "").strip()
                if not asset_id:
                    raise HTTPException(status_code=400, detail="Missing asset_id for server item.")
                cand = (base_autofill / asset_id).resolve()
                try:
                    cand.relative_to(base_autofill.resolve())
                except ValueError as e:
                    raise HTTPException(status_code=400, detail="Invalid asset_id path.") from e
                if not cand.is_file():
                    raise HTTPException(status_code=404, detail=f"Server asset not found: {asset_id}")
                media_paths.append(cand)
                captions.append(caption)
                titles.append(title)
                caption_texts.append(ct_raw)
                overlay_positions.append((float(overlay_x), float(overlay_y)))
                overlay_font_scales.append(float(font_scale))
            elif src == "upload":
                upload_index = it.get("upload_index")
                if upload_index is None:
                    raise HTTPException(status_code=400, detail="Missing upload_index for upload item.")
                if not isinstance(upload_index, int):
                    raise HTTPException(status_code=400, detail="upload_index must be an integer.")
                if upload_index < 0 or upload_index >= len(uploaded_paths):
                    raise HTTPException(status_code=400, detail="upload_index out of range.")
                media_paths.append(uploaded_paths[upload_index])
                captions.append(caption)
                titles.append(title)
                caption_texts.append(ct_raw)
                overlay_positions.append((float(overlay_x), float(overlay_y)))
                overlay_font_scales.append(float(font_scale))
            elif src == "url":
                raw_url = str(it.get("url") or "").strip()
                if not raw_url:
                    raise HTTPException(status_code=400, detail="Missing url for url item.")
                mtype = str(it.get("media_type") or it.get("type") or "image").lower().strip()
                if mtype not in ("image", "video"):
                    mtype = "image"
                digest = hashlib.sha256(raw_url.encode("utf-8")).hexdigest()[:12]

                if raw_url.startswith("/media/"):
                    local_src = _resolve_local_media_path(raw_url)
                    if local_src is None or not local_src.is_file():
                        raise HTTPException(
                            status_code=404,
                            detail=f"Local media not found (save the session first or check path): {raw_url[:160]}",
                        )
                    ext = local_src.suffix.lower() or (".mp4" if mtype == "video" else ".jpg")
                    name = f"local_{len(media_paths):02d}_{digest}{ext}"
                    dest = temp_dir / name
                    await asyncio.to_thread(shutil.copy2, local_src, dest)
                elif raw_url.startswith(("https://", "http://")):
                    base = raw_url.split("?", 1)[0].lower()
                    if mtype == "video":
                        ext = ".mp4"
                        for cand in (".mp4", ".webm", ".mov", ".m4v"):
                            if base.endswith(cand):
                                ext = cand
                                break
                    else:
                        ext = ".jpg"
                        for cand in (".jpg", ".jpeg", ".png", ".webp"):
                            if base.endswith(cand):
                                ext = cand
                                break
                    lib_hit = await asyncio.to_thread(_find_ad_reels_library_local_path, raw_url)
                    if lib_hit is not None:
                        lext = lib_hit.suffix.lower()
                        if lext in (
                            ".jpg",
                            ".jpeg",
                            ".png",
                            ".webp",
                            ".mp4",
                            ".webm",
                            ".mov",
                            ".m4v",
                        ):
                            ext = lext
                    name = f"url_{len(media_paths):02d}_{digest}{ext}"
                    dest = temp_dir / name
                    if lib_hit is not None:
                        await asyncio.to_thread(shutil.copy2, lib_hit, dest)
                    else:
                        try:
                            await asyncio.to_thread(
                                media_processor.download_binary,
                                raw_url,
                                dest,
                            )
                        except Exception as e:
                            raise HTTPException(
                                status_code=502,
                                detail=f"Failed to download media URL: {e}",
                            ) from e
                    if not dest.is_file() or dest.stat().st_size == 0:
                        raise HTTPException(
                            status_code=502,
                            detail="Media file missing or empty after copy/download.",
                        )
                else:
                    raise HTTPException(
                        status_code=400,
                        detail="url must be http(s), or /media/... pointing at a file under output.",
                    )
                media_paths.append(dest)
                captions.append(caption)
                titles.append(title)
                caption_texts.append(ct_raw)
                overlay_positions.append((float(overlay_x), float(overlay_y)))
                overlay_font_scales.append(float(font_scale))
            else:
                raise HTTPException(status_code=400, detail=f"Unknown item source: {src!r}")
    else:
        if not files:
            raise HTTPException(status_code=400, detail="Upload at least one file or provide items_json.")
        try:
            parsed = json.loads(captions_json or "[]")
            captions = [str(x or "") for x in parsed] if isinstance(parsed, list) else []
        except json.JSONDecodeError as e:
            raise HTTPException(status_code=400, detail="Invalid captions_json payload.") from e

        media_paths = list(uploaded_paths)
        # Legacy mode: keep previous default bottom-center caption placement.
        overlay_positions = [manual_reel_builder.DEFAULT_OVERLAY_ANCHOR] * len(media_paths)
        overlay_font_scales = [requested_font_scale] * len(media_paths)

    # User-controlled anchor (0–1): center of the caption block on the frame.
    overlay_positions = [(oax, oay)] * len(media_paths)
    overlay_font_scales = [requested_font_scale] * len(media_paths)

    try:
        if not media_paths:
            raise HTTPException(status_code=400, detail="No media items supplied.")

        res = await asyncio.to_thread(
            manual_reel_builder.build_manual_reel,
            uploads_dir=temp_dir,
            media_paths=media_paths,
            captions=captions,
            music_track_id=music_track_id,
            transition_type=transition_type,
            transition_speed=transition_speed,
            transition_xfade_scale=xfade_scale_opt,
            overlay_positions=overlay_positions,
            overlay_font_scales=overlay_font_scales,
            titles=titles if items_json else None,
            caption_texts=caption_texts if items_json else None,
            hook_caption=hook_caption,
            hook_seconds=hook_sec_f,
            image_segment_seconds=clip_img,
            video_segment_seconds=clip_vid,
            show_branding=show_brand_on_reel,
            music_volume=max(0.01, min(2.0, float(music_volume or "0.3"))),
        )
        out = dict(res)
        out["video_url"] = _to_media_url(out.get("output_path") or "") or None
        return JSONResponse(content=out)
    except HTTPException:
        raise
    except RuntimeError as e:
        raise HTTPException(status_code=502, detail=str(e)) from e
    except Exception:
        logger.exception("Upload reel generation failed")
        raise HTTPException(status_code=500, detail="Upload reel generation failed") from None


@app.post("/api/upload-reel/convert-landscape")
async def api_upload_reel_convert_landscape(
    video_url: str = Form(...),
    music_volume: str = Form(default="0.3"),
) -> JSONResponse:
    """
    Convert a 9:16 reel to 16:9 landscape (YouTube format).

    Uses a blurred + zoomed copy of the source as background, with the
    original portrait video centred on top — the standard YouTube
    Shorts-to-landscape conversion style.
    """
    import subprocess as _sp

    rel = video_url.lstrip("/")
    src = config.OUTPUT_DIR.parent / rel if not rel.startswith("output/") else config.OUTPUT_DIR.parent / rel
    if not src.is_file():
        src = config.OUTPUT_DIR / Path(rel).relative_to("media") if rel.startswith("media/") else None  # type: ignore[assignment]
        if src is None or not src.is_file():  # type: ignore[union-attr]
            raise HTTPException(status_code=404, detail=f"Source video not found: {video_url}")

    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    out_dir = config.OUTPUT_DIR / "youtube_reels"
    out_dir.mkdir(parents=True, exist_ok=True)
    slug = re.sub(r"[^a-z0-9]+", "-", src.stem.lower())[:40]
    out_mp4 = out_dir / f"{ts}_{slug}_16x9.mp4"

    exe = media_processor._ensure_ffmpeg()  # type: ignore[attr-defined]
    vol = max(0.01, min(2.0, float(music_volume or "0.3")))

    # Blurred background: scale source to fill 1920x1080 + heavy blur
    # Foreground: scale source to fit height=1080, centre it
    filter_complex = (
        "[0:v]scale=1920:1080:force_original_aspect_ratio=increase,"
        "crop=1920:1080,boxblur=30:5[bg];"
        "[0:v]scale=-2:1080[fg];"
        "[bg][fg]overlay=(W-w)/2:(H-h)/2[v]"
    )
    cmd = [
        exe, "-y",
        "-i", str(src),
        "-filter_complex", filter_complex,
        "-map", "[v]",
        "-map", "0:a?",
        "-c:v", "libx264", "-crf", "23", "-preset", "fast",
        "-c:a", "aac", "-b:a", "192k",
        "-af", f"volume={vol}",
        "-movflags", "+faststart",
        str(out_mp4),
    ]

    def _run() -> None:
        media_processor._run_ffmpeg_cmd(cmd, "landscape convert")  # type: ignore[attr-defined]

    try:
        await asyncio.to_thread(_run)
    except RuntimeError as e:
        raise HTTPException(status_code=502, detail=str(e)) from e

    video_url_out = "/media/" + out_mp4.relative_to(config.OUTPUT_DIR).as_posix()
    return JSONResponse({"ok": True, "video_url": video_url_out, "format": "16x9", "path": str(out_mp4)})


@app.post("/api/upload-reel/autofill")
async def api_upload_reel_autofill(body: AutofillReelMediaBody) -> JSONResponse:
    theme = (body.theme or "").strip()
    if not theme:
        raise HTTPException(status_code=400, detail="theme is required.")

    try:
        res = await asyncio.to_thread(
            manual_reel_autofill.autofill_media_for_theme,
            theme,
            max_items=body.max_items,
            include_video=body.include_video,
        )
        return JSONResponse(content=res)
    except RuntimeError as e:
        raise HTTPException(status_code=502, detail=str(e)) from e
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Autofill failed")
        raise HTTPException(status_code=500, detail=str(e)) from e


def _carousel_slides_media_urls(run_dir: Path) -> list[str]:
    """Web URLs for existing carousel slide JPEGs under this run (original 9:16 assets)."""
    sj = run_dir / "summary.json"
    if not sj.is_file():
        return []
    try:
        data = json.loads(sj.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return []
    outputs = data.get("outputs") or {}
    slides = outputs.get("carousel_slides") or []
    if not isinstance(slides, list):
        return []
    urls: list[str] = []
    for s in slides:
        p = Path(str(s))
        if not p.is_file():
            continue
        u = _to_media_url(p)
        if u:
            urls.append(u)
    return urls


@app.get("/carousel/{run_id}", response_class=HTMLResponse)
async def carousel_run_page(request: Request, run_id: str) -> HTMLResponse:
    """All original carousel slide images for one pipeline run."""
    config.ensure_output_dirs()
    try:
        run_dir = safe_carousel_run_dir(run_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e
    if not run_dir.is_dir():
        raise HTTPException(status_code=404, detail="Run not found.")
    sj = run_dir / "summary.json"
    if not sj.is_file():
        raise HTTPException(status_code=404, detail="No summary.json for this run.")
    try:
        data = json.loads(sj.read_text(encoding="utf-8"))
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=404, detail="Invalid summary.json.") from e
    content = data.get("content") or {}
    theme = (data.get("theme") or "").strip()
    hook = (content.get("hook") or "").strip()
    slides_urls = _carousel_slides_media_urls(run_dir)
    page_title = hook or theme or "Carousel slides"
    return templates.TemplateResponse(
        "carousel_run.html",
        {
            "request": request,
            "title": page_title,
            "nav_active": "reels",
            "run_id": run_id,
            "theme": theme,
            "carousel_slides_urls": slides_urls,
            "slide_count": len(slides_urls),
        },
    )


@app.get("/api/reels")
async def api_reels() -> JSONResponse:
    """All reels from ``carousel/*/summary.json`` (newest first)."""
    config.ensure_output_dirs()
    return JSONResponse(content={"reels": reels_catalog.list_reels()})


@app.get("/api/reels/export.xlsx")
async def api_reels_export_xlsx() -> StreamingResponse:
    """Excel workbook of the reel library."""
    config.ensure_output_dirs()
    rows = reels_catalog.list_reels()
    wb = Workbook()
    ws = wb.active
    ws.title = "Reels"
    headers = [
        "Run ID",
        "Theme",
        "Title (hook)",
        "Hashtags",
        "Generated (UTC)",
        "Reel filename",
        "Reel file present",
        "Reel path (web)",
    ]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=r.get("run_id") or "")
        ws.cell(row=i, column=2, value=r.get("theme") or "")
        ws.cell(row=i, column=3, value=r.get("title") or "")
        ws.cell(row=i, column=4, value=r.get("hashtags") or "")
        ws.cell(row=i, column=5, value=r.get("generated_at") or "")
        ws.cell(row=i, column=6, value=r.get("reel_filename") or "")
        ws.cell(row=i, column=7, value="yes" if r.get("reel_exists") else "no")
        ws.cell(row=i, column=8, value=r.get("reel_video_url") or "")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"velo_reels_{stamp}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/api/generate")
async def api_generate(body: GenerateBody) -> JSONResponse:
    theme = body.theme.strip()
    if not theme:
        raise HTTPException(status_code=400, detail="Theme is required.")

    try:
        summary = await asyncio.to_thread(
            pipeline.run_pipeline,
            theme,
            body.music_track_id,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e
    except RuntimeError as e:
        logger.exception("Pipeline failed")
        raise HTTPException(status_code=502, detail=str(e)) from e
    except Exception as e:
        logger.exception("Unexpected pipeline error")
        raise HTTPException(status_code=500, detail=str(e)) from e

    enriched = _enrich_summary_for_web(summary)
    return JSONResponse(content=enriched)


@app.get("/api/music-tracks")
async def api_music_tracks() -> JSONResponse:
    """Filenames under ``music/`` for the background-music dropdown."""
    config.ensure_output_dirs()
    return JSONResponse(content={"tracks": config.list_music_tracks()})


@app.get("/api/music/auto-track")
async def api_music_auto_track(mood: str = Query(default="travel")) -> JSONResponse:
    """
    Return a music track_id for use in reel generation.

    Priority:
      1. Jamendo API (if JAMENDO_CLIENT_ID is set) — fetches a popular
         royalty-free track matching ``mood``, caches the MP3 under music/jamendo/.
      2. Local music/ folder — picks a random file.
      3. No music (returns track_id=null).
    """
    import random

    jamendo_id = config.JAMENDO_CLIENT_ID
    jamendo_cache_dir = config.MUSIC_LIBRARY_DIR / "jamendo"

    if jamendo_id:
        try:
            # Map mood to Jamendo search terms (instrumental tracks only)
            search_map = {
                "travel": "cinematic",
                "cinematic": "cinematic",
                "upbeat": "upbeat",
                "relaxed": "relaxing",
                "adventure": "adventure",
                "inspiring": "inspiring",
            }
            search_term = search_map.get(mood.lower(), "cinematic")

            async with httpx.AsyncClient(timeout=15) as client:
                resp = await client.get(
                    "https://api.jamendo.com/v3.0/tracks/",
                    params={
                        "client_id": jamendo_id,
                        "format": "json",
                        "limit": 20,
                        "search": search_term,
                        "vocalinstrumental": "instrumental",
                        "audiodlformat": "mp32",
                        "order": "popularity_total",
                    },
                )
                resp.raise_for_status()
                data = resp.json()

            results = data.get("results") or []
            if results:
                track = random.choice(results[:10])  # pick from top-10 most popular
                dl_url = track.get("audiodownload") or track.get("audio")
                track_id_str = str(track.get("id", ""))
                title = track.get("name", "")
                artist = track.get("artist_name", "")

                if dl_url and track_id_str:
                    jamendo_cache_dir.mkdir(parents=True, exist_ok=True)
                    local_path = jamendo_cache_dir / f"{track_id_str}.mp3"

                    if not local_path.exists():
                        async with httpx.AsyncClient(timeout=60, follow_redirects=True) as dl:
                            audio = await dl.get(dl_url)
                            audio.raise_for_status()
                            local_path.write_bytes(audio.content)

                    rel_id = local_path.relative_to(config.MUSIC_LIBRARY_DIR).as_posix()
                    return JSONResponse({
                        "source": "jamendo",
                        "track_id": rel_id,
                        "title": title,
                        "artist": artist,
                        "jamendo_id": track_id_str,
                    })
        except Exception as exc:
            logger.warning("Jamendo fetch failed: %s — falling back to local music", exc)

    # Fallback: random local track
    tracks = config.list_music_tracks()
    if tracks:
        import random as _r
        t = _r.choice(tracks)
        return JSONResponse({"source": "local", "track_id": t["id"], "title": t["label"]})

    return JSONResponse({"source": "none", "track_id": None, "title": ""})


@app.get("/api/health")
async def health(debug: bool = Query(default=False)) -> dict[str, Any]:
    """Use ``?debug=1`` to see which ``web_app.py`` is loaded and whether ``/ad-reels`` is registered."""
    out: dict[str, Any] = {"status": "ok"}
    if debug:
        paths = {getattr(r, "path", None) for r in app.routes if getattr(r, "path", None)}
        out["web_app_path"] = str(Path(__file__).resolve())
        out["has_ad_reels"] = "/ad-reels" in paths
    return out


if __name__ == "__main__":
    import sys

    import uvicorn

    _root = Path(__file__).resolve().parent.parent
    sys.path.insert(0, str(_root))
    uvicorn.run(
        "velo_web:app",
        host="127.0.0.1",
        port=8000,
        reload=True,
    )
